#!/usr/bin/env python3
"""
Build a self-contained Excel workbook for Webinar → Application Pipeline Analysis.
La Trobe University brand theme. Apple-style pill design on Summary tab.

Run this ONCE on your personal machine. The generated .xlsx file works entirely
with Excel formulas and native charts — no scripts needed on the work computer.

Usage:
    python3 build_workbook.py

Output:
    Flexi_PG_Webinar_Analysis_2024_25.xlsx
"""

import os
import random
import copy
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.drawing.image import Image as XlImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection as CellProtection
from openpyxl.comments import Comment

# ─── Configuration ──────────────────────────────────────────────────────
NUM_WEBINAR = 40
NUM_FLEXI = 30
NUM_OVERLAP = 15
DATA_START_ROW = 2
MAX_DATA_ROW = 5000
OUTPUT_FILE = "Flexi_PG_Webinar_Analysis_2024_25.xlsx"
LOGO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "latrobe_header.png")

# ─── La Trobe University brand palette ──────────────────────────────────
CHERRY    = "E4002B"   # primary cherry red
DARK_RED  = "A6192E"   # deep red
CHARCOAL  = "2D2926"   # charcoal
WARM_GREY = "6D6E71"   # warm grey
LT_TEAL   = "00857C"   # accent teal
GREEN     = "27AE60"
LIGHT_BG  = "FDF6F6"   # warm off-white
WHITE     = "FFFFFF"
GREY      = "D1CCBD"   # sandstone
DARK_GREY = "6D6E71"

# Apple-style pill fills
PILL_GREEN  = "D5F5E3"
PILL_RED    = "FADBD8"
PILL_BLUE   = "D6EAF8"
PILL_AMBER  = "FEF3E2"
PILL_PURPLE = "E8DAEF"
PILL_GREY   = "F2F3F4"

# ─── Style helpers ──────────────────────────────────────────────────────
header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=CHARCOAL, end_color=CHARCOAL, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color=GREY),
    right=Side(style="thin", color=GREY),
    top=Side(style="thin", color=GREY),
    bottom=Side(style="thin", color=GREY),
)


def style_header_row(ws, num_cols, row=1):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 30
    ws.auto_filter.ref = f"A{row}:{get_column_letter(num_cols)}{row}"


def alt_row_shading(ws, start_row, end_row, num_cols):
    light = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).fill = light


def add_logo(ws, anchor_cell="A1"):
    """Add La Trobe University logo image to a sheet at given anchor."""
    if os.path.exists(LOGO_FILE):
        logo = XlImage(LOGO_FILE)
        logo.width = 150   # ~2 inches in Excel
        logo.height = 79
        ws.add_image(logo, anchor_cell)


def ltu_banner(ws, end_col_letter="L", row=1):
    """Cherry-red La Trobe banner across the top of a sheet."""
    ws.merge_cells(f"A{row}:{end_col_letter}{row}")
    cell = ws.cell(row=row, column=1,
                   value="          La Trobe University")  # indented for logo
    cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=CHERRY, end_color=CHERRY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 55  # taller for logo
    # Fill all cells in the merge range with the cherry colour
    from openpyxl.utils import column_index_from_string
    for c in range(1, column_index_from_string(end_col_letter) + 1):
        ws.cell(row=row, column=c).fill = PatternFill(
            start_color=CHERRY, end_color=CHERRY, fill_type="solid")
    # Add logo overlaid on the banner
    add_logo(ws, f"A{row}")


# ─── Fictional data generators ──────────────────────────────────────────
FIRST_NAMES = [
    "Emma", "Liam", "Olivia", "Noah", "Ava", "Ethan", "Sophia", "Mason",
    "Isabella", "James", "Mia", "Benjamin", "Charlotte", "Lucas", "Amelia",
    "Henry", "Harper", "Alexander", "Evelyn", "Daniel", "Abigail", "Matthew",
    "Emily", "Jack", "Scarlett", "Owen", "Grace", "Sebastian", "Chloe",
    "Aiden", "Lily", "William", "Zoe", "Elijah", "Riley", "Logan",
    "Aria", "Jackson", "Penelope", "Leo", "Nora", "Gabriel", "Hannah",
    "Samuel", "Ella", "David", "Hazel", "Joseph", "Aurora", "Carter"
]
LAST_NAMES = [
    "Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller",
    "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez",
    "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin",
    "Lee", "Perez", "Thompson", "White", "Harris", "Sanchez", "Clark",
    "Ramirez", "Lewis", "Robinson", "Walker", "Young", "Allen", "King",
    "Wright", "Scott", "Torres", "Nguyen", "Hill", "Flores", "Green",
    "Adams", "Nelson", "Baker", "Hall", "Rivera", "Campbell", "Mitchell"
]
COURSES = [
    "Master of Business Administration",
    "Master of Data Science",
    "Master of Information Technology",
    "Master of Professional Accounting",
    "Master of Cybersecurity",
    "Master of Engineering Management",
    "Master of Public Health",
    "Master of Biotechnology and Bioinformatics",
    "Master of Health Information Management",
    "Master of Financial Analysis",
    "Master of Nursing Practice",
]
COURSE_CODES = [
    "LMBAF", "LMDSC", "LMITF", "LMPAF", "LMCYB",
    "LMENM", "LMPHE", "LMBBI", "LMHIM", "LMFAN", "LMNUP"
]
CAMPUSES = ["Melbourne (Bundoora)", "Sydney", "Online", "City (Melbourne)"]
STATUSES = ["Attended", "Attended", "Attended", "No Show", "No Show",
            "Registered", "Attended", "Attended"]
DESCRIBES = [
    "Working professional", "Recent graduate", "Career changer",
    "International student", "Current student", "Parent returning to study"
]
AGE_RANGES = ["18-24", "25-34", "35-44", "45-54", "55+"]
LOCATIONS_AU = [
    "Melbourne, VIC", "Sydney, NSW", "Brisbane, QLD", "Perth, WA",
    "Adelaide, SA", "Canberra, ACT", "Hobart, TAS", "Darwin, NT",
    "Gold Coast, QLD", "Geelong, VIC"
]
HOW_FOUND = [
    "Social Media", "Google Search", "Word of Mouth", "University Website",
    "Email Newsletter", "Education Fair", "LinkedIn", "Friend/Family"
]
FEE_TYPES = ["Domestic", "International"]
SCHOOL_OWNERS = [
    "La Trobe Business School",
    "School of Computing, Engineering and Mathematical Sciences",
    "School of Nursing and Midwifery",
    "School of Allied Health, Human Services and Sport",
    "School of Agriculture, Biomedicine and Environment",
]

# Course → School lookup for the hidden SchoolMap sheet
SCHOOL_MAP = {
    "Master of Business Administration": "La Trobe Business School",
    "Master of Professional Accounting": "La Trobe Business School",
    "Master of Financial Analysis": "La Trobe Business School",
    "Master of Data Science": "School of Computing, Engineering and Mathematical Sciences",
    "Master of Information Technology": "School of Computing, Engineering and Mathematical Sciences",
    "Master of Cybersecurity": "School of Computing, Engineering and Mathematical Sciences",
    "Master of Engineering Management": "School of Computing, Engineering and Mathematical Sciences",
    "Master of Public Health": "School of Allied Health, Human Services and Sport",
    "Master of Biotechnology and Bioinformatics": "School of Agriculture, Biomedicine and Environment",
    "Master of Health Information Management": "School of Allied Health, Human Services and Sport",
    "Master of Nursing Practice": "School of Nursing and Midwifery",
}
SCHOOLS = sorted(set(SCHOOL_MAP.values()))


def rand_email(first, last):
    domains = ["gmail.com", "outlook.com", "yahoo.com", "hotmail.com", "icloud.com"]
    return f"{first.lower()}.{last.lower()}{random.randint(1,99)}@{random.choice(domains)}"


def rand_phone():
    return f"04{random.randint(10000000, 99999999)}"


def rand_date(start, end):
    return start + timedelta(days=random.randint(0, (end - start).days))


# ═══════════════════════════════════════════════════════════════════════
#  Apple-style pill helpers for Summary sheet
# ═══════════════════════════════════════════════════════════════════════
def pill_cell(ws, row, col, value, fill_hex, font_color="333333",
              bold=False, size=11):
    """Style a cell to look like an Apple-design pill/tag."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=font_color)
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex,
                            fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color=fill_hex),
        right=Side(style="thin", color=fill_hex),
        top=Side(style="thin", color=fill_hex),
        bottom=Side(style="thin", color=fill_hex),
    )
    return cell


def pill_value(ws, row, col, formula, fill_hex="F8F9FA", font_color="2D2926",
               bold=True, size=13):
    """Style a formula value cell as a pill with prominent number."""
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    cell.font = Font(name="Calibri", size=size, bold=bold, color=font_color)
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex,
                            fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin", color="E8E8E8"),
        right=Side(style="thin", color="E8E8E8"),
        top=Side(style="thin", color="E8E8E8"),
        bottom=Side(style="thin", color="E8E8E8"),
    )
    return cell


def section_title(ws, row, col, title, end_col=3):
    """Apple-style section header — clean, left-aligned, charcoal."""
    for c in range(col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=CHARCOAL, end_color=CHARCOAL,
                                fill_type="solid")
        cell.border = Border(
            bottom=Side(style="medium", color=CHERRY))
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28


def insight_q(ws, row, col, question):
    """Concise insight question — italic grey, explains what the section answers.
    Placed in spacer rows so it annotates the section immediately below."""
    cell = ws.cell(row=row, column=col, value=question)
    cell.font = Font(name="Calibri", size=10, italic=True, color="999999")
    cell.alignment = Alignment(horizontal="left", vertical="bottom")


# ═══════════════════════════════════════════════════════════════════════
#  BUILD
# ═══════════════════════════════════════════════════════════════════════
def build():
    wb = Workbook()

    # ================================================================
    # SHEET 1 — Instructions
    # ================================================================
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.sheet_properties.tabColor = CHERRY

    # La Trobe banner
    ltu_banner(ws_inst, "A")
    ws_inst.column_dimensions["A"].width = 85

    instructions = [
        ("", None),
        ("Webinar → Enrolment Pipeline Analysis", "title"),
        ("La Trobe University — Postgraduate Webinar Pipeline", "subtitle"),
        ("", None),
        ("HOW TO USE THIS WORKBOOK", "section"),
        ("", None),
        ("1. Go to the 'Webinar data 2024_25' sheet (green tab)", None),
        ("   - Delete the sample data rows (keep the header row!)", None),
        ("   - Paste your real webinar data starting from row 3", None),
        ("   - First column is Webinar Year, then STATUS, etc.", None),
        ("", None),
        ("2. Go to the 'Flexi PG 2024_25' sheet (green tab)", None),
        ("   - Delete the sample data rows (keep the header row!)", None),
        ("   - Paste your real applicant data starting from row 3", None),
        ("   - Column AD = Offers count, Column AG = Commencing Enrolments", None),
        ("", None),
        ("3. The 'Analysis' sheet has formula-driven columns:", None),
        ("   - Columns A–Q: Registration, Attendance, Application, Pipeline", None),
        ("   - Column R: Webinar_Year (copied from source data)", None),
        ("   - Column S: School (inferred from Flexi PG or course preference)", None),
        ("   - Column T: Has_Offer (from Flexi PG column AD)", None),
        ("   - Column U: Has_Enrolled (from Flexi PG column AG)", None),
        ("   - Column V: Extended_Pipeline (full 5-stage status)", None),
        ("   - Column W: Student_Type (Domestic or International)", None),
        ("   - These recalculate automatically — DO NOT edit", None),
        ("", None),
        ("4. The 'Summary' sheet shows aggregated counts", None),
        ("   - Apple-style pill cards for quick scanning", None),
        ("   - Enrolment Pipeline section (Registered → Enrolled)", None),
        ("   - Domestic vs International breakdown", None),
        ("   - School Breakdown with webinar counts per school", None),
        ("   - Raw values alongside for copy-paste", None),
        ("", None),
        ("5. The 'Dashboard' sheet has KPIs, charts, and findings", None),
        ("   - 8 KPI cards: Total, Attended, Applied, Conversion,", None),
        ("     Offered, Offer Rate, Enrolled, Enrolment Rate", None),
        ("   - 15 charts including Enrolment Funnel, School, Student Type,", None),
        ("     Offer Acceptance, Yield Composition, and App-to-Offer Rate", None),
        ("   - 15 auto-generated findings with commentary", None),
        ("   - Everything updates when data changes", None),
        ("", None),
        ("FILTERS (on Dashboard)", "section"),
        ("- On the Dashboard sheet, row 4 has two dropdowns:", None),
        ("  * Year: All / 2024 / 2025 (cell E4)", None),
        ("  * Student Type: All / Domestic / International (cell I4)", None),
        ("- These filters control ALL data across Summary and Dashboard", None),
        ("- The Summary sheet mirrors the current filter values (read-only)", None),
        ("- 'Domestic' = Australian citizen or permanent resident (column L = 'Yes')", None),
        ("- 'International' = all others (column L = 'No')", None),
        ("- Combine both filters freely, e.g. Year=2025 + Student=Domestic", None),
        ("", None),
        ("PIPELINE FUNNEL", "section"),
        ("- Shows Registered → Attended → Applied → Offered → Enrolled", None),
        ("- '% of Total' = percentage of all registrants at each stage", None),
        ("- 'Stage→Stage' = conversion from previous stage (e.g. 30% of attendees applied)", None),
        ("", None),
        ("NO-SHOW CONVERSION", "section"),
        ("- Shows what % of non-attendees still applied, received offers, and enrolled", None),
        ("- Answers the question: 'Did not attend but still applied and got an offer'", None),
        ("- KPI cards on Dashboard show No-Show→Applied % and No-Show→Offered %", None),
        ("", None),
        ("APPLICATION-TO-OFFER RATE", "section"),
        ("- Shows how many applicants received offers, split by attendance path", None),
        ("- Answers: 'Are webinar attendees stronger candidates than no-shows?'", None),
        ("- Chart 15 on Dashboard visualises this as a clustered bar chart", None),
        ("", None),
        ("SCHOOL BREAKDOWN", "section"),
        ("- Schools are inferred from Flexi PG data or course preferences", None),
        ("- A hidden 'SchoolMap' sheet maps courses to La Trobe schools", None),
        ("- To add new courses: unprotect SchoolMap (password: view),", None),
        ("  add the course name in column A and school in column B", None),
        ("", None),
        ("IMPORTANT NOTES", "section"),
        ("- Formulas cover up to row 5000 — add more rows if needed", None),
        ("- Green tabs = your data (editable)", None),
        ("- Other tabs = formulas/charts (protected, password: view)", None),
        ("- Tab order: Instructions → Dashboard → Summary → Analysis → Data", None),
        ("", None),
        (f"Generated: {datetime.now().strftime('%d %b %Y at %H:%M')}", None),
    ]

    title_font = Font(name="Calibri", size=20, bold=True, color=CHERRY)
    subtitle_font = Font(name="Calibri", size=12, italic=True, color=WARM_GREY)
    section_font = Font(name="Calibri", size=13, bold=True, color=DARK_RED)
    body_font = Font(name="Calibri", size=11, color=CHARCOAL)

    for i, (text, style) in enumerate(instructions, 2):
        cell = ws_inst.cell(row=i, column=1, value=text)
        if style == "title":
            cell.font = title_font
        elif style == "subtitle":
            cell.font = subtitle_font
        elif style == "section":
            cell.font = section_font
        else:
            cell.font = body_font

    # ================================================================
    # SHEET 2 — Webinar data 2024_25
    # ================================================================
    ws_web = wb.create_sheet("Webinar data 2024_25")
    ws_web.sheet_properties.tabColor = "27AE60"  # green = editable

    # Column layout (A=Webinar Year is NEW):
    # A: Webinar Year    B: STATUS           C: Engagement score
    # D: Minutes         E: First name       F: Last name
    # G: Email           H: Mobile phone     I: What best describes you?
    # J: Age range       K: Location         L: Are you a citizen or PR?
    # M: Are you residing in Australia?       N: Course Preference
    # O: Campus interest P: How did you find out?
    web_headers = [
        "Webinar Year",
        "STATUS", "Engagement score", "Minutes", "First name", "Last name",
        "Email", "Mobile phone", "What best describes you?", "Age range",
        "Location", "Are you a citizen or PR?",
        "Are you residing in Australia?", "Course Preference (course name)",
        "Campus interest", "How did you find out about the event?"
    ]

    end_col_letter = get_column_letter(len(web_headers))
    ltu_banner(ws_web, end_col_letter)

    for c, h in enumerate(web_headers, 1):
        ws_web.cell(row=2, column=c, value=h)
    style_header_row(ws_web, len(web_headers), row=2)

    # Shared identities for overlap
    shared_people = []
    for _ in range(NUM_OVERLAP):
        fn = random.choice(FIRST_NAMES)
        ln = random.choice(LAST_NAMES)
        shared_people.append((fn, ln, rand_email(fn, ln), rand_phone()))

    webinar_rows = []
    for i in range(NUM_WEBINAR):
        if i < NUM_OVERLAP:
            fn, ln, email, phone = shared_people[i]
        else:
            fn, ln = random.choice(FIRST_NAMES), random.choice(LAST_NAMES)
            email, phone = rand_email(fn, ln), rand_phone()

        status = random.choice(STATUSES)
        eng = random.randint(10, 100) if status == "Attended" else 0
        mins = random.randint(5, 90) if status == "Attended" else 0
        course_idx = random.randint(0, len(COURSES) - 1)

        webinar_rows.append([
            random.choice(["2024", "2025"]),
            status, eng, mins, fn, ln, email, phone,
            random.choice(DESCRIBES), random.choice(AGE_RANGES),
            random.choice(LOCATIONS_AU),
            random.choice(["Yes", "No"]), random.choice(["Yes", "No"]),
            COURSES[course_idx], random.choice(CAMPUSES),
            random.choice(HOW_FOUND),
        ])

    for r_idx, row in enumerate(webinar_rows, 3):  # data starts row 3
        for c_idx, val in enumerate(row, 1):
            ws_web.cell(row=r_idx, column=c_idx, value=val)

    alt_row_shading(ws_web, 3, 3 + NUM_WEBINAR - 1, len(web_headers))
    for c in range(1, len(web_headers) + 1):
        ws_web.column_dimensions[get_column_letter(c)].width = 20

    # ================================================================
    # SHEET 3 — Flexi PG 2024_25
    # ================================================================
    ws_flexi = wb.create_sheet("Flexi PG 2024_25")
    ws_flexi.sheet_properties.tabColor = "27AE60"

    flexi_headers = [
        "Date Created", "Surname", "Given Name", "Given Name 2",
        "Email", "Mobile Phone",
        "LTU Course Code", "LTU Course Version", "LTU Campus",
        "LTU Availability Number", "LTU Course Full-Title",
        "Fee Type", "Pref_Order", "Offered_Date", "Quota_Managed_Crs",
        "Liability Course Category Code",
        "Liability Course Category Description",
        "Enrolment Year", "Intake Number", "Intake",
        "Course Category Group", "Course Category Type",
        "Full-Course Title", "School Course Owner",
        "Risepoint Course Indicator.ORG_UNIT_NM",
        "Application Liability Category Code",
        "Application Liability Category",
        "Applications", "Conditional Offers", "Offers",
        "Acceptances", "Enrolments", "Commencing Enrolments"
    ]
    flexi_end_col = get_column_letter(len(flexi_headers))
    ltu_banner(ws_flexi, flexi_end_col)

    for c, h in enumerate(flexi_headers, 1):
        ws_flexi.cell(row=2, column=c, value=h)
    style_header_row(ws_flexi, len(flexi_headers), row=2)

    flexi_rows = []
    for i in range(NUM_FLEXI):
        if i < NUM_OVERLAP:
            fn, ln, email, phone = shared_people[i]
            mt = random.choice(["both", "both", "email_only", "phone_only"])
            if mt == "email_only":
                phone = rand_phone()
            elif mt == "phone_only":
                email = rand_email(fn, ln) + ".au"
        else:
            fn, ln = random.choice(FIRST_NAMES), random.choice(LAST_NAMES)
            email, phone = rand_email(fn, ln), rand_phone()

        ci = random.randint(0, len(COURSES) - 1)
        dc = rand_date(datetime(2024, 1, 1), datetime(2025, 6, 30))
        apps = random.randint(1, 3)
        cond = random.randint(0, apps)
        off = random.randint(0, cond)
        acc = random.randint(0, off)
        enr = random.randint(0, acc)

        flexi_rows.append([
            dc.strftime("%d/%m/%Y"), ln, fn, "",
            email, phone,
            COURSE_CODES[ci], "1", random.choice(CAMPUSES),
            str(random.randint(1000, 9999)), COURSES[ci],
            random.choice(FEE_TYPES), str(random.randint(1, 3)),
            (dc + timedelta(days=random.randint(14, 60))).strftime("%d/%m/%Y"),
            random.choice(["Y", "N"]),
            f"LC{random.randint(100,999)}",
            random.choice(["Postgraduate Coursework", "Postgraduate Research"]),
            str(dc.year), str(random.randint(1, 3)),
            random.choice(["Semester 1", "Semester 2",
                           "Trimester 1", "Trimester 2", "Trimester 3"]),
            "Postgraduate", "Coursework", COURSES[ci],
            random.choice(SCHOOL_OWNERS), random.choice(["N/A", "Risepoint"]),
            f"ALC{random.randint(100,999)}",
            random.choice(["Domestic Fee-Paying", "International", "CSP"]),
            apps, cond, off, acc, enr, random.randint(0, enr),
        ])

    for r_idx, row in enumerate(flexi_rows, 3):
        for c_idx, val in enumerate(row, 1):
            ws_flexi.cell(row=r_idx, column=c_idx, value=val)

    alt_row_shading(ws_flexi, 3, 3 + NUM_FLEXI - 1, len(flexi_headers))
    for c in range(1, len(flexi_headers) + 1):
        ws_flexi.column_dimensions[get_column_letter(c)].width = 18

    # ================================================================
    # SHEET 3b — SchoolMap  (hidden lookup table)
    # ================================================================
    ws_map = wb.create_sheet("SchoolMap")
    ws_map.cell(row=1, column=1, value="Course Name").font = Font(bold=True)
    ws_map.cell(row=1, column=2, value="School").font = Font(bold=True)
    for i, (course, school) in enumerate(SCHOOL_MAP.items(), 2):
        ws_map.cell(row=i, column=1, value=course)
        ws_map.cell(row=i, column=2, value=school)
    ws_map.column_dimensions["A"].width = 48
    ws_map.column_dimensions["B"].width = 55
    ws_map.sheet_state = "hidden"
    ws_map.protection.sheet = True
    ws_map.protection.password = "view"

    # ================================================================
    # SHEET 4 — Analysis  (formula-driven)
    # ================================================================
    # With Webinar Year as col A, the new Webinar sheet column map:
    #   A=Webinar Year  B=STATUS  E=First name  F=Last name
    #   G=Email  H=Mobile phone  J=Age range  K=Location
    #   N=Course Preference  O=Campus interest  P=How did you find out
    # Data starts row 3 in webinar sheet (row 1=banner, row 2=headers)

    ws_anal = wb.create_sheet("Analysis")
    ws_anal.sheet_properties.tabColor = CHERRY

    anal_headers = [
        "Row#",                   # A
        "First Name",             # B
        "Last Name",              # C
        "Email",                  # D
        "Phone",                  # E
        "STATUS",                 # F
        "Attendance_Category",    # G
        "Email_Match",            # H
        "Phone_Match",            # I
        "Has_Applied",            # J
        "Match_Method",           # K
        "Course_Preference",      # L
        "Campus_Interest",        # M
        "Age_Range",              # N
        "Location",               # O
        "Discovery_Channel",      # P
        "Pipeline_Stage",         # Q
        "Webinar_Year",           # R
        "School",                 # S
        "Has_Offer",              # T
        "Has_Enrolled",           # U
        "Extended_Pipeline",      # V
        "Student_Type",           # W
    ]
    ltu_banner(ws_anal, "W")
    for c, h in enumerate(anal_headers, 1):
        ws_anal.cell(row=2, column=c, value=h)
    style_header_row(ws_anal, len(anal_headers), row=2)

    # Formulas: row 3..MAX_DATA_ROW+2  (data in webinar sheet also starts row 3)
    for r in range(3, MAX_DATA_ROW + 3):
        wr = r  # same row in webinar sheet

        # Existence check: webinar col E (First name)
        chk = f"'Webinar data 2024_25'!E{wr}"

        # A: Row#
        ws_anal.cell(row=r, column=1).value = \
            f'=IF({chk}="","",ROW()-2)'

        # B: First Name  (webinar col E)
        ws_anal.cell(row=r, column=2).value = \
            f'=IF({chk}="","",{chk})'

        # C: Last Name  (webinar col F)
        ws_anal.cell(row=r, column=3).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!F{wr})'

        # D: Email  (webinar col G)
        ws_anal.cell(row=r, column=4).value = \
            f'=IF({chk}="","",LOWER(TRIM(\'Webinar data 2024_25\'!G{wr})))'

        # E: Phone  (webinar col H)
        ws_anal.cell(row=r, column=5).value = \
            f'=IF({chk}="","",TRIM(\'Webinar data 2024_25\'!H{wr}))'

        # F: STATUS  (webinar col B)
        ws_anal.cell(row=r, column=6).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!B{wr})'

        # G: Attendance_Category
        ws_anal.cell(row=r, column=7).value = \
            f'=IF(F{r}="","",IF(LOWER(TRIM(F{r}))="attended","Registered and Attended","Registered and Not Attended"))'

        # H: Email_Match  (Flexi PG col E = Email)
        ws_anal.cell(row=r, column=8).value = \
            f'=IF(D{r}="","",IF(COUNTIF(\'Flexi PG 2024_25\'!E:E,D{r})>0,1,0))'

        # I: Phone_Match  (Flexi PG col F = Mobile Phone)
        ws_anal.cell(row=r, column=9).value = \
            f'=IF(E{r}="","",IF(COUNTIF(\'Flexi PG 2024_25\'!F:F,E{r})>0,1,0))'

        # J: Has_Applied
        ws_anal.cell(row=r, column=10).value = \
            f'=IF(D{r}="","",IF(OR(H{r}=1,I{r}=1),"Applied","Did Not Apply"))'

        # K: Match_Method
        ws_anal.cell(row=r, column=11).value = (
            f'=IF(D{r}="","",IF(AND(H{r}=1,I{r}=1),"Email & Phone",'
            f'IF(H{r}=1,"Email Only",IF(I{r}=1,"Phone Only","No Match"))))'
        )

        # L: Course Preference  (webinar col N)
        ws_anal.cell(row=r, column=12).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!N{wr})'

        # M: Campus Interest  (webinar col O)
        ws_anal.cell(row=r, column=13).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!O{wr})'

        # N: Age Range  (webinar col J — &"" forces text coercion for COUNTIFS)
        ws_anal.cell(row=r, column=14).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!J{wr}&"")'

        # O: Location  (webinar col K)
        ws_anal.cell(row=r, column=15).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!K{wr})'

        # P: Discovery Channel  (webinar col P — &"" forces text coercion for COUNTIFS)
        ws_anal.cell(row=r, column=16).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!P{wr}&"")'

        # Q: Pipeline Stage
        ws_anal.cell(row=r, column=17).value = (
            f'=IF(D{r}="","",IF(J{r}="Did Not Apply",'
            f'IF(G{r}="Registered and Attended","Attended - No Application","Registered Only"),'
            f'"Webinar \u2192 Applicant"))'
        )

        # R: Webinar_Year  (webinar col A — +0 forces NUMBER output so
        #    the dropdown value (also NUMBER) compares correctly in
        #    COUNTIFS and SUMPRODUCT — NUMBER=NUMBER, no type mismatch)
        ws_anal.cell(row=r, column=18).value = \
            f'=IF({chk}="","",\'Webinar data 2024_25\'!A{wr}+0)'

        # S: School — try email match → phone match → VLOOKUP SchoolMap → "Unknown"
        #    &"" on INDEX/VLOOKUP results forces text coercion for COUNTIFS
        ws_anal.cell(row=r, column=19).value = (
            f'=IF(D{r}="","",IF(H{r}=1,'
            f'IFERROR(INDEX(\'Flexi PG 2024_25\'!X:X,MATCH(D{r},\'Flexi PG 2024_25\'!E:E,0))&"","Unknown"),'
            f'IF(I{r}=1,'
            f'IFERROR(INDEX(\'Flexi PG 2024_25\'!X:X,MATCH(E{r},\'Flexi PG 2024_25\'!F:F,0))&"","Unknown"),'
            f'IFERROR(VLOOKUP(L{r},SchoolMap!A:B,2,FALSE)&"","Unknown"))))'
        )

        # T: Has_Offer — if Applied, check Flexi PG col AD (Offers) > 0
        ws_anal.cell(row=r, column=20).value = (
            f'=IF(D{r}="","",IF(J{r}<>"Applied","N/A",'
            f'IF(H{r}=1,'
            f'IF(SUMIFS(\'Flexi PG 2024_25\'!AD:AD,\'Flexi PG 2024_25\'!E:E,D{r})>0,"Offered","Not Offered"),'
            f'IF(SUMIFS(\'Flexi PG 2024_25\'!AD:AD,\'Flexi PG 2024_25\'!F:F,E{r})>0,"Offered","Not Offered"))))'
        )

        # U: Has_Enrolled — if Applied, check Flexi PG col AG (Commencing Enrolments) > 0
        ws_anal.cell(row=r, column=21).value = (
            f'=IF(D{r}="","",IF(J{r}<>"Applied","N/A",'
            f'IF(H{r}=1,'
            f'IF(SUMIFS(\'Flexi PG 2024_25\'!AG:AG,\'Flexi PG 2024_25\'!E:E,D{r})>0,"Enrolled","Not Enrolled"),'
            f'IF(SUMIFS(\'Flexi PG 2024_25\'!AG:AG,\'Flexi PG 2024_25\'!F:F,E{r})>0,"Enrolled","Not Enrolled"))))'
        )

        # V: Extended_Pipeline — full funnel stage
        ws_anal.cell(row=r, column=22).value = (
            f'=IF(D{r}="","",IF(U{r}="Enrolled","Enrolled",'
            f'IF(T{r}="Offered","Offered - Not Enrolled",'
            f'IF(J{r}="Applied","Applied - No Offer",'
            f'IF(G{r}="Registered and Attended","Attended - No Application","Registered Only")))))'
        )

        # W: Student_Type — Domestic (citizen/PR) vs International
        # &"" forces string coercion; LEFT(…,1)="Y" avoids direct
        # cross-sheet full-string comparison (Mac Excel inlineStr bug).
        ws_anal.cell(row=r, column=23).value = (
            f'=IF(LEN(\'Webinar data 2024_25\'!A{wr}&"")=0,"",'
            f'IF(LEFT(\'Webinar data 2024_25\'!L{wr}&"",1)="Y","Domestic","International"))'
        )

    for c in range(1, len(anal_headers) + 1):
        ws_anal.column_dimensions[get_column_letter(c)].width = 22

    # Conditional formatting — pill-like colours for key columns
    gf = PatternFill(start_color=PILL_GREEN, end_color=PILL_GREEN, fill_type="solid")
    rf = PatternFill(start_color=PILL_RED, end_color=PILL_RED, fill_type="solid")
    bf = PatternFill(start_color=PILL_BLUE, end_color=PILL_BLUE, fill_type="solid")
    af = PatternFill(start_color=PILL_AMBER, end_color=PILL_AMBER, fill_type="solid")

    MR = MAX_DATA_ROW + 2
    # Has_Applied (col J)
    ws_anal.conditional_formatting.add(
        f"J3:J{MR}", CellIsRule(operator="equal", formula=['"Applied"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"J3:J{MR}", CellIsRule(operator="equal", formula=['"Did Not Apply"'], fill=rf))
    # Attendance_Category (col G)
    ws_anal.conditional_formatting.add(
        f"G3:G{MR}", CellIsRule(operator="equal",
                                formula=['"Registered and Attended"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"G3:G{MR}", CellIsRule(operator="equal",
                                formula=['"Registered and Not Attended"'], fill=rf))
    # Match_Method (col K)
    ws_anal.conditional_formatting.add(
        f"K3:K{MR}", CellIsRule(operator="equal",
                                formula=['"Email & Phone"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"K3:K{MR}", CellIsRule(operator="equal",
                                formula=['"Email Only"'], fill=bf))
    ws_anal.conditional_formatting.add(
        f"K3:K{MR}", CellIsRule(operator="equal",
                                formula=['"Phone Only"'], fill=af))
    # Pipeline_Stage (col Q)
    ws_anal.conditional_formatting.add(
        f"Q3:Q{MR}", CellIsRule(operator="equal",
                                formula=['"Webinar \u2192 Applicant"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"Q3:Q{MR}", CellIsRule(operator="equal",
                                formula=['"Registered Only"'], fill=rf))
    # Has_Offer (col T)
    ws_anal.conditional_formatting.add(
        f"T3:T{MR}", CellIsRule(operator="equal", formula=['"Offered"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"T3:T{MR}", CellIsRule(operator="equal", formula=['"Not Offered"'], fill=rf))
    # Has_Enrolled (col U)
    ws_anal.conditional_formatting.add(
        f"U3:U{MR}", CellIsRule(operator="equal", formula=['"Enrolled"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"U3:U{MR}", CellIsRule(operator="equal", formula=['"Not Enrolled"'], fill=rf))
    # Extended_Pipeline (col V)
    ws_anal.conditional_formatting.add(
        f"V3:V{MR}", CellIsRule(operator="equal", formula=['"Enrolled"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"V3:V{MR}", CellIsRule(operator="equal", formula=['"Registered Only"'], fill=rf))
    ws_anal.conditional_formatting.add(
        f"V3:V{MR}", CellIsRule(operator="equal",
                                formula=['"Attended - No Application"'], fill=af))
    # Student_Type (col W)
    ws_anal.conditional_formatting.add(
        f"W3:W{MR}", CellIsRule(operator="equal", formula=['"Domestic"'], fill=gf))
    ws_anal.conditional_formatting.add(
        f"W3:W{MR}", CellIsRule(operator="equal", formula=['"International"'], fill=af))

    # ================================================================
    # SHEET 5 — Summary  (Apple-style pill design)
    # ================================================================
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = DARK_RED

    ltu_banner(ws_sum, "E")

    # Column widths — generous spacing for pill layout
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 16
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 16

    ws_sum.sheet_view.showGridLines = False  # clean Apple look

    row_h = 26  # pill row height

    # ─── Filter Display (row 2) — mirrors Dashboard filters (read-only) ─────
    ws_sum.cell(row=2, column=1, value="Year:").font = Font(
        name="Calibri", size=11, bold=True, color=CHARCOAL)
    ws_sum.cell(row=2, column=1).alignment = Alignment(
        horizontal="right", vertical="center")
    filter_cell = ws_sum.cell(row=2, column=2)
    filter_cell.value = "=Dashboard!E4"  # mirrors Dashboard year filter
    filter_cell.font = Font(name="Calibri", size=11, bold=True, color=CHERRY)
    filter_cell.alignment = Alignment(horizontal="center", vertical="center")
    filter_cell.border = Border(
        left=Side(style="thin", color=CHERRY),
        right=Side(style="thin", color=CHERRY),
        top=Side(style="thin", color=CHERRY),
        bottom=Side(style="thin", color=CHERRY))

    ws_sum.cell(row=2, column=4, value="Student:").font = Font(
        name="Calibri", size=11, bold=True, color=CHARCOAL)
    ws_sum.cell(row=2, column=4).alignment = Alignment(
        horizontal="right", vertical="center")
    st_cell = ws_sum.cell(row=2, column=5)
    st_cell.value = "=Dashboard!I4"  # mirrors Dashboard student filter
    st_cell.font = Font(name="Calibri", size=11, bold=True, color=CHERRY)
    st_cell.alignment = Alignment(horizontal="center", vertical="center")
    st_cell.border = Border(
        left=Side(style="thin", color=CHERRY),
        right=Side(style="thin", color=CHERRY),
        top=Side(style="thin", color=CHERRY),
        bottom=Side(style="thin", color=CHERRY))

    # Light background for filter row — makes it visually distinct
    filter_bg = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    for fc in range(1, 7):  # A2:F2
        c = ws_sum.cell(row=2, column=fc)
        if c.fill == PatternFill():  # only fill if not already styled
            c.fill = filter_bg
    ws_sum.row_dimensions[2].height = 28

    # Analysis range references (data starts row 3 in Analysis)
    A_G = f"Analysis!G3:G{MR}"   # Attendance_Category
    A_J = f"Analysis!J3:J{MR}"   # Has_Applied
    A_K = f"Analysis!K3:K{MR}"   # Match_Method
    A_Q = f"Analysis!Q3:Q{MR}"   # Pipeline_Stage
    A_D = f"Analysis!D3:D{MR}"   # Email (for total count)
    A_N = f"Analysis!N3:N{MR}"   # Age Range
    A_M = f"Analysis!M3:M{MR}"   # Campus Interest
    A_P = f"Analysis!P3:P{MR}"   # Discovery Channel
    A_R = f"Analysis!R3:R{MR}"   # Webinar_Year
    A_S = f"Analysis!S3:S{MR}"   # School
    A_T = f"Analysis!T3:T{MR}"   # Has_Offer
    A_U = f"Analysis!U3:U{MR}"   # Has_Enrolled
    A_V = f"Analysis!V3:V{MR}"   # Extended_Pipeline
    A_W = f"Analysis!W3:W{MR}"   # Student_Type

    # Year filter shorthand — used in every formula
    # Column R stores NUMBERS (via +0). The dropdown may give TEXT "2024" or
    # NUMBER 2024 — we can't control Excel's behaviour. The *1 on the criteria
    # forces it to NUMBER regardless, so the comparison is always NUMBER=NUMBER.
    # When B2="All", the IF short-circuits to "<>" / 1, so *1 never runs on "All".
    YF = 'Dashboard!$E$4'                        # year filter cell (Dashboard)
    YC = f'{A_R},IF({YF}="All","<>",{YF}*1)'             # year criterion for COUNTIFS
    # (YM SUMPRODUCT multiplier removed — broke on Mac Excel; all formulas
    #  now use the COUNTIFS-based YC instead.)

    # Student type filter shorthand — appended after year filter
    SF = 'Dashboard!$I$4'                        # student filter cell (Dashboard)
    SC = f'{A_W},IF({SF}="All","<>",{SF})'       # student criterion for COUNTIFS
    # (SM SUMPRODUCT multiplier removed — same Mac Excel issue as YM.)

    # Fully-filtered total registrants (year + student type)
    # COUNTIFS with "?*" (matches any non-empty cell) replaces the old
    # SUMPRODUCT(--(LEN(D)>0)*YM*SM) which broke on Mac Excel because
    # SUMPRODUCT array comparisons against formula-returned / cross-sheet
    # values don't evaluate correctly.  COUNTIFS handles them fine.
    TOTAL_F = f'COUNTIFS({A_D},"?*",{YC},{SC})'

    # ─── KPI Pills (row 3-5) ──────────────────
    r = 3
    section_title(ws_sum, r, 1, "KEY METRICS", end_col=6)
    r += 1
    ws_sum.row_dimensions[r].height = 20  # spacer

    r += 1  # row 5 — KPI labels
    kpi_pills = [
        ("Total\nRegistrants", PILL_GREY, f'={TOTAL_F}'),
        ("Attended", PILL_GREEN, f'=COUNTIFS({A_G},"Registered and Attended",{YC},{SC})'),
        ("Not Attended", PILL_RED, f'=COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})'),
        ("Applied", PILL_BLUE, f'=COUNTIFS({A_J},"Applied",{YC},{SC})'),
        ("Conversion %", PILL_AMBER,
         f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_J},"Applied",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")'),
    ]

    for i, (label, fill, formula) in enumerate(kpi_pills):
        col = i + 1
        pill_cell(ws_sum, r, col, label, fill, font_color=CHARCOAL,
                  bold=True, size=9)
        ws_sum.row_dimensions[r].height = 32
        pill_value(ws_sum, r + 1, col, formula, fill_hex=WHITE,
                   font_color=CHERRY, bold=True, size=18)
        ws_sum.row_dimensions[r + 1].height = 40

    # Dynamic filter context — row 7 shows active filter + zero warning
    ws_sum.merge_cells("A7:F7")
    ctx = ws_sum.cell(row=7, column=1)
    ctx.value = (
        f'=IF(A6=0,'
        f'"⚠ No registrants match: Year="&{YF}&", Student="&{SF}'
        f'&" — try changing filters",'
        f'"Filtered: "&IF({YF}="All","All years",{YF})'
        f'&" · "&IF({SF}="All","All students",{SF})'
        f'&" — "&TEXT(A6,"#,##0")&" registrants")'
    )
    ctx.font = Font(name="Calibri", size=10, bold=True, color="E67E22")
    ctx.alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[7].height = 20

    # Dynamic insight — adapts to filter state
    dyn_insight = ws_sum.cell(row=8, column=1)
    dyn_insight.value = (
        f'=IF(A6=0,'
        f'"↳ This filter combination has no data — "'
        f'&IF({SF}<>"All","the webinar data column L (citizen/PR) controls Domestic/International","check the Year dropdown"),'
        f'"↳ Did registrants actually show up? "&TEXT(B6,"#,##0")&" attended out of "&TEXT(A6,"#,##0"))'
    )
    dyn_insight.font = Font(name="Calibri", size=10, italic=True, color="999999")
    dyn_insight.alignment = Alignment(horizontal="left", vertical="bottom")

    # ─── Attendance Breakdown (row 9) ──────────
    r = 9
    section_title(ws_sum, r, 1, "ATTENDANCE BREAKDOWN", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "Category", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 2, "Count", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "% Share", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    r += 1
    pill_cell(ws_sum, r, 1, "Registered and Attended", PILL_GREEN,
              font_color="1B7A3D", bold=False, size=10)
    pill_value(ws_sum, r, 2, f'=COUNTIFS({A_G},"Registered and Attended",{YC},{SC})',
              fill_hex=PILL_GREEN, font_color="1B7A3D", size=12)
    pill_value(ws_sum, r, 3,
              f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Attended",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
              fill_hex=PILL_GREEN, font_color="1B7A3D", size=12)
    ws_sum.row_dimensions[r].height = row_h

    r += 1
    pill_cell(ws_sum, r, 1, "Registered and Not Attended", PILL_RED,
              font_color="922B21", bold=False, size=10)
    pill_value(ws_sum, r, 2, f'=COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})',
              fill_hex=PILL_RED, font_color="922B21", size=12)
    pill_value(ws_sum, r, 3,
              f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
              fill_hex=PILL_RED, font_color="922B21", size=12)
    ws_sum.row_dimensions[r].height = row_h

    # Dynamic: shows applied count when data exists, warns on zero
    dyn13 = ws_sum.cell(row=13, column=1)
    dyn13.value = (
        f'=IF(A6=0,"↳ No data — adjust filters above",'
        f'"↳ Did attendance convert into applications? "'
        f'&TEXT(COUNTIFS({A_J},"Applied",{YC},{SC}),"#,##0")&" applied")'
    )
    dyn13.font = Font(name="Calibri", size=10, italic=True, color="999999")
    dyn13.alignment = Alignment(horizontal="left", vertical="bottom")

    # ─── Application Status (row 14) ──────────
    r = 14
    section_title(ws_sum, r, 1, "APPLICATION STATUS", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "Status", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 2, "Count", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "% Share", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    r += 1  # 16
    pill_cell(ws_sum, r, 1, "Applied", PILL_GREEN, font_color="1B7A3D", size=10)
    pill_value(ws_sum, r, 2, f'=COUNTIFS({A_J},"Applied",{YC},{SC})',
              fill_hex=PILL_GREEN, font_color="1B7A3D", size=12)
    pill_value(ws_sum, r, 3,
              f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_J},"Applied",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
              fill_hex=PILL_GREEN, font_color="1B7A3D", size=12)
    ws_sum.row_dimensions[r].height = row_h

    r += 1  # 17
    pill_cell(ws_sum, r, 1, "Did Not Apply", PILL_RED, font_color="922B21", size=10)
    pill_value(ws_sum, r, 2, f'=COUNTIFS({A_J},"Did Not Apply",{YC},{SC})',
              fill_hex=PILL_RED, font_color="922B21", size=12)
    pill_value(ws_sum, r, 3,
              f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_J},"Did Not Apply",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
              fill_hex=PILL_RED, font_color="922B21", size=12)
    ws_sum.row_dimensions[r].height = row_h

    insight_q(ws_sum, 19, 1, "↳ Does attending actually improve application rates?")

    # ─── Cross-tab: Attendance × Application (row 35) ──
    r = 35
    section_title(ws_sum, r, 1, "ATTENDANCE \u00d7 APPLICATION CROSS-TAB", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "", CHARCOAL, font_color=WHITE, size=10)
    pill_cell(ws_sum, r, 2, "Applied", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "Did Not Apply", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    r += 1  # 37
    pill_cell(ws_sum, r, 1, "Attended", PILL_GREEN, font_color="1B7A3D", size=10)
    pill_value(ws_sum, r, 2,
              f'=COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})',
              fill_hex=PILL_GREEN, font_color="1B7A3D", size=12)
    pill_value(ws_sum, r, 3,
              f'=COUNTIFS({A_G},"Registered and Attended",{A_J},"Did Not Apply",{YC},{SC})',
              fill_hex=PILL_AMBER, font_color="7D6608", size=12)
    ws_sum.row_dimensions[r].height = row_h

    r += 1  # 38
    pill_cell(ws_sum, r, 1, "Not Attended", PILL_RED, font_color="922B21", size=10)
    pill_value(ws_sum, r, 2,
              f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})',
              fill_hex=PILL_BLUE, font_color="21618C", size=12)
    pill_value(ws_sum, r, 3,
              f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Did Not Apply",{YC},{SC})',
              fill_hex=PILL_RED, font_color="922B21", size=12)
    ws_sum.row_dimensions[r].height = row_h

    insight_q(ws_sum, 40, 1, "↳ Where should we focus marketing spend?")

    # ─── Discovery Channel (row 41) ───────────
    r = 41
    section_title(ws_sum, r, 1, "DISCOVERY CHANNEL", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "Channel", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 2, "Total", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "Applied", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    for i, ch in enumerate(HOW_FOUND):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, ch, PILL_GREY, font_color=CHARCOAL, size=10)
        pill_value(ws_sum, rr, 2, f'=COUNTIFS({A_P},"{ch}",{YC},{SC})',
                  fill_hex=PILL_GREY, font_color=CHARCOAL, size=11)
        pill_value(ws_sum, rr, 3,
                  f'=COUNTIFS({A_P},"{ch}",{A_J},"Applied",{YC},{SC})',
                  fill_hex=PILL_GREEN, font_color="1B7A3D", size=11)
        ws_sum.row_dimensions[rr].height = row_h

    discovery_end_row = r + len(HOW_FOUND)
    insight_q(ws_sum, discovery_end_row + 1, 1, "↳ Which age groups show the most interest?")

    # ─── Age Range (after discovery) ──────────
    r = discovery_end_row + 2
    age_start = r
    section_title(ws_sum, r, 1, "AGE RANGE", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "Age", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 2, "Total", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "Applied", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    for i, ar in enumerate(AGE_RANGES):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, ar, PILL_PURPLE, font_color="6C3483", size=10)
        pill_value(ws_sum, rr, 2, f'=COUNTIFS({A_N},"{ar}",{YC},{SC})',
                  fill_hex=PILL_PURPLE, font_color="6C3483", size=11)
        pill_value(ws_sum, rr, 3,
                  f'=COUNTIFS({A_N},"{ar}",{A_J},"Applied",{YC},{SC})',
                  fill_hex=PILL_GREEN, font_color="1B7A3D", size=11)
        ws_sum.row_dimensions[rr].height = row_h

    age_end_row = r + len(AGE_RANGES)
    insight_q(ws_sum, age_end_row + 1, 1, "↳ Which campus draws the most applicants?")

    # ─── Campus (after age) ───────────────────
    r = age_end_row + 2
    campus_start = r
    section_title(ws_sum, r, 1, "CAMPUS INTEREST", end_col=6)
    r += 1
    pill_cell(ws_sum, r, 1, "Campus", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 2, "Total", CHARCOAL, font_color=WHITE, bold=True, size=10)
    pill_cell(ws_sum, r, 3, "Applied", CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    for i, cp in enumerate(CAMPUSES):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, cp, PILL_BLUE, font_color="21618C", size=10)
        pill_value(ws_sum, rr, 2, f'=COUNTIFS({A_M},"{cp}",{YC},{SC})',
                  fill_hex=PILL_BLUE, font_color="21618C", size=11)
        pill_value(ws_sum, rr, 3,
                  f'=COUNTIFS({A_M},"{cp}",{A_J},"Applied",{YC},{SC})',
                  fill_hex=PILL_GREEN, font_color="1B7A3D", size=11)
        ws_sum.row_dimensions[rr].height = row_h

    campus_end_row = r + len(CAMPUSES)
    # Dynamic: when filtered to a specific student type, explain it's redundant
    dyn_stud = ws_sum.cell(row=campus_end_row + 1, column=1)
    dyn_stud.value = (
        f'=IF({SF}<>"All",'
        f'"↳ You\'re filtering to "&{SF}&" only — this section reflects that single group",'
        f'"↳ What\'s the domestic vs international split?")'
    )
    dyn_stud.font = Font(name="Calibri", size=10, italic=True, color="999999")
    dyn_stud.alignment = Alignment(horizontal="left", vertical="bottom")

    # ─── Domestic vs International (after campus) ───
    r = campus_end_row + 2
    section_title(ws_sum, r, 1, "DOMESTIC vs INTERNATIONAL", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Type", "Total", "Attended", "Applied", "Enrolled", "% of Total"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.column_dimensions["F"].width = 16  # match col E
    ws_sum.row_dimensions[r].height = row_h

    for i, stype in enumerate(["Domestic", "International"]):
        rr = r + 1 + i
        fc = "1B7A3D" if stype == "Domestic" else "7D6608"
        fill = PILL_GREEN if stype == "Domestic" else PILL_AMBER
        pill_cell(ws_sum, rr, 1, stype, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2,
                  f'=COUNTIFS({A_W},"{stype}",{YC},{SC})',
                  fill_hex=fill, font_color=fc, size=11)
        pill_value(ws_sum, rr, 3,
                  f'=COUNTIFS({A_W},"{stype}",{A_G},"Registered and Attended",{YC},{SC})',
                  fill_hex=fill, font_color=fc, size=11)
        pill_value(ws_sum, rr, 4,
                  f'=COUNTIFS({A_W},"{stype}",{A_J},"Applied",{YC},{SC})',
                  fill_hex=fill, font_color=fc, size=11)
        pill_value(ws_sum, rr, 5,
                  f'=COUNTIFS({A_W},"{stype}",{A_U},"Enrolled",{YC},{SC})',
                  fill_hex=fill, font_color=fc, size=11)
        pill_value(ws_sum, rr, 6,
                  f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_W},"{stype}",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
                  fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    student_end_row = r + 2  # 2 types
    insight_q(ws_sum, student_end_row + 1, 1, "↳ Where do we lose students in the funnel?")

    # ─── Pipeline Funnel (after student type) ─────────────────────
    # Implements the manager's hand-drawn flow diagram:
    #   Webinar → Registered → Attended? → Applied? → Offered? → Enrolled?
    # Each stage shows:  Count | % of Total | Stage→Stage conversion
    # The Stage→Stage column answers "of those who reached stage N,
    # what % moved to stage N+1?" — the manager's "30% of attendees applied"
    r = student_end_row + 2
    pipeline_start = r
    section_title(ws_sum, r, 1, "PIPELINE FUNNEL", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Stage", "Count", "% of Total", "Stage\u2192Stage"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    # Shorthand formulas for each stage count (reused across columns)
    _REG = TOTAL_F
    _ATT = f'COUNTIFS({A_G},"Registered and Attended",{YC},{SC})'
    _APP = f'COUNTIFS({A_J},"Applied",{YC},{SC})'
    _OFF = f'COUNTIFS({A_T},"Offered",{YC},{SC})'
    _ENR = f'COUNTIFS({A_U},"Enrolled",{YC},{SC})'

    # (label, fill, fontcolor, count_formula, pct_of_total, stage_conversion)
    funnel_stages = [
        ("Registered", PILL_GREY, CHARCOAL,
         f'={_REG}',
         f'="100%"',
         f'="\u2014"'),  # baseline — no previous stage
        ("Attended", PILL_GREEN, "1B7A3D",
         f'={_ATT}',
         f'=IF({_REG}=0,0,TEXT(ROUND({_ATT}/{_REG}*100,1),"0.0")&"%")',
         f'=IF({_REG}=0,0,TEXT(ROUND({_ATT}/{_REG}*100,1),"0.0")&"%")'),
        ("Applied", PILL_BLUE, "21618C",
         f'={_APP}',
         f'=IF({_REG}=0,0,TEXT(ROUND({_APP}/{_REG}*100,1),"0.0")&"%")',
         f'=IF({_ATT}=0,0,TEXT(ROUND({_APP}/{_ATT}*100,1),"0.0")&"%")'),
        ("Offered", PILL_AMBER, "7D6608",
         f'={_OFF}',
         f'=IF({_REG}=0,0,TEXT(ROUND({_OFF}/{_REG}*100,1),"0.0")&"%")',
         f'=IF({_APP}=0,0,TEXT(ROUND({_OFF}/{_APP}*100,1),"0.0")&"%")'),
        ("Enrolled", PILL_GREEN, "1B7A3D",
         f'={_ENR}',
         f'=IF({_REG}=0,0,TEXT(ROUND({_ENR}/{_REG}*100,1),"0.0")&"%")',
         f'=IF({_OFF}=0,0,TEXT(ROUND({_ENR}/{_OFF}*100,1),"0.0")&"%")'),
    ]
    for i, (label, fill, fc, count_f, pct_f, conv_f) in enumerate(funnel_stages):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, label, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2, count_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 3, pct_f, fill_hex=fill, font_color=fc, size=11)
        pill_value(ws_sum, rr, 4, conv_f, fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    pipeline_end_row = r + len(funnel_stages)
    insight_q(ws_sum, pipeline_end_row + 1, 1, "↳ Are no-shows a complete loss?")

    # ─── No-Show Conversion (after pipeline funnel) ──────────────
    # Manager's request: registrants who did NOT attend but still
    # applied / got offered / enrolled — shown as % of all no-shows.
    r = pipeline_end_row + 2
    section_title(ws_sum, r, 1, "NO-SHOW CONVERSION", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Metric", "Count", "% of No-Shows"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    _NOSHOW = f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})'

    noshow_rows = [
        ("Did Not Attend", PILL_GREY, CHARCOAL,
         f'={_NOSHOW}',
         f'="\u2014"'),
        ("Still Applied", PILL_BLUE, "21618C",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})',
         f'=IF({_NOSHOW}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})/{_NOSHOW}*100,1),"0.0")&"%")'),
        ("Received Offer", PILL_AMBER, "7D6608",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})',
         f'=IF({_NOSHOW}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})/{_NOSHOW}*100,1),"0.0")&"%")'),
        ("Enrolled", PILL_GREEN, "1B7A3D",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_U},"Enrolled",{YC},{SC})',
         f'=IF({_NOSHOW}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_U},"Enrolled",{YC},{SC})/{_NOSHOW}*100,1),"0.0")&"%")'),
    ]
    for i, (label, fill, fc, count_f, pct_f) in enumerate(noshow_rows):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, label, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2, count_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 3, pct_f, fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    noshow_end_row = r + len(noshow_rows)
    insight_q(ws_sum, noshow_end_row + 1, 1, "↳ Once offered, do students actually enrol?")

    # ─── Offer Acceptance (after no-show conversion) ─────────────
    r = noshow_end_row + 2
    offer_start = r
    section_title(ws_sum, r, 1, "OFFER ACCEPTANCE", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Metric", "Count", "Rate"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    # Shorthand for offer-acceptance formulas
    _TOT_OFF = f'COUNTIFS({A_T},"Offered",{YC},{SC})'
    _TOT_ENR = f'COUNTIFS({A_T},"Offered",{A_U},"Enrolled",{YC},{SC})'
    _ATT_OFF = f'COUNTIFS({A_G},"Registered and Attended",{A_T},"Offered",{YC},{SC})'
    _ATT_ENR = f'COUNTIFS({A_G},"Registered and Attended",{A_T},"Offered",{A_U},"Enrolled",{YC},{SC})'
    _NS_OFF  = f'COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})'
    _NS_ENR  = f'COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{A_U},"Enrolled",{YC},{SC})'

    offer_rows = [
        ("Total Offered", PILL_GREY, CHARCOAL,
         f'={_TOT_OFF}',
         f'="\u2014"'),
        ("Accepted (Enrolled)", PILL_GREEN, "1B7A3D",
         f'={_TOT_ENR}',
         f'=IF({_TOT_OFF}=0,0,TEXT(ROUND({_TOT_ENR}/{_TOT_OFF}*100,1),"0.0")&"%")'),
        ("Not Accepted", PILL_AMBER, "7D6608",
         f'={_TOT_OFF}-{_TOT_ENR}',
         f'=IF({_TOT_OFF}=0,0,TEXT(ROUND(({_TOT_OFF}-{_TOT_ENR})/{_TOT_OFF}*100,1),"0.0")&"%")'),
        ("Attended \u2192 Accepted", PILL_BLUE, "21618C",
         f'={_ATT_ENR}',
         f'=IF({_ATT_OFF}=0,0,TEXT(ROUND({_ATT_ENR}/{_ATT_OFF}*100,1),"0.0")&"%")'),
        ("No-Show \u2192 Accepted", PILL_BLUE, "21618C",
         f'={_NS_ENR}',
         f'=IF({_NS_OFF}=0,0,TEXT(ROUND({_NS_ENR}/{_NS_OFF}*100,1),"0.0")&"%")'),
    ]
    for i, (label, fill, fc, count_f, pct_f) in enumerate(offer_rows):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, label, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2, count_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 3, pct_f, fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    offer_end_row = r + len(offer_rows)
    insight_q(ws_sum, offer_end_row + 1, 1, "↳ Which schools are driving the pipeline?")

    # ─── School Breakdown (after offer acceptance) ───
    r = offer_end_row + 2
    school_start = r
    section_title(ws_sum, r, 1, "SCHOOL BREAKDOWN", end_col=6)
    r += 1
    for ci, hdr in enumerate(["School", "Registered", "Attended", "Applied", "Enrolled"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    all_schools = SCHOOLS + ["Unknown"]
    for i, school in enumerate(all_schools):
        rr = r + 1 + i
        short_name = school if len(school) <= 30 else school[:28] + ".."
        pill_cell(ws_sum, rr, 1, short_name, PILL_GREY, font_color=CHARCOAL, size=9)
        pill_value(ws_sum, rr, 2,
                  f'=COUNTIFS({A_S},"{school}",{YC},{SC})',
                  fill_hex=PILL_GREY, font_color=CHARCOAL, size=11)
        pill_value(ws_sum, rr, 3,
                  f'=COUNTIFS({A_S},"{school}",{A_G},"Registered and Attended",{YC},{SC})',
                  fill_hex=PILL_GREEN, font_color="1B7A3D", size=11)
        pill_value(ws_sum, rr, 4,
                  f'=COUNTIFS({A_S},"{school}",{A_J},"Applied",{YC},{SC})',
                  fill_hex=PILL_BLUE, font_color="21618C", size=11)
        pill_value(ws_sum, rr, 5,
                  f'=COUNTIFS({A_S},"{school}",{A_U},"Enrolled",{YC},{SC})',
                  fill_hex=PILL_GREEN, font_color="1B7A3D", size=11)
        ws_sum.row_dimensions[rr].height = row_h

    school_end_row = r + len(all_schools)
    insight_q(ws_sum, school_end_row + 1, 1, "↳ Is the webinar driving enrolment, or do people enrol regardless?")

    # ─── Yield Composition (after school breakdown) ───────────────
    # Manager question: "Are enrolled students coming from the webinar,
    # or would they have enrolled anyway?" Split enrolled by attendance path.
    r = school_end_row + 2
    yield_start = r
    section_title(ws_sum, r, 1, "YIELD COMPOSITION", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Metric", "Count", "% of Enrolled"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    _YIELD_ENR = f'COUNTIFS({A_U},"Enrolled",{YC},{SC})'
    _ATT_ENR_Y = f'COUNTIFS({A_G},"Registered and Attended",{A_U},"Enrolled",{YC},{SC})'
    _NS_ENR_Y  = f'COUNTIFS({A_G},"Registered and Not Attended",{A_U},"Enrolled",{YC},{SC})'

    yield_rows = [
        ("Total Enrolled", PILL_GREY, CHARCOAL,
         f'={_YIELD_ENR}',
         f'="\u2014"'),
        ("Attended \u2192 Enrolled", PILL_GREEN, "1B7A3D",
         f'={_ATT_ENR_Y}',
         f'=IF({_YIELD_ENR}=0,0,TEXT(ROUND({_ATT_ENR_Y}/{_YIELD_ENR}*100,1),"0.0")&"%")'),
        ("No-Show \u2192 Enrolled", PILL_AMBER, "7D6608",
         f'={_NS_ENR_Y}',
         f'=IF({_YIELD_ENR}=0,0,TEXT(ROUND({_NS_ENR_Y}/{_YIELD_ENR}*100,1),"0.0")&"%")'),
    ]
    for i, (label, fill, fc, count_f, pct_f) in enumerate(yield_rows):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, label, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2, count_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 3, pct_f, fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    yield_end_row = r + len(yield_rows)
    insight_q(ws_sum, yield_end_row + 1, 1, "↳ Are webinar-sourced applicants strong candidates?")

    # ─── Application-to-Offer Rate (after yield composition) ──────
    # Compare offer rates for attended vs no-show applicants.
    # If attended applicants get offers at a higher rate, the webinar
    # is producing stronger candidates.
    r = yield_end_row + 2
    a2o_start = r
    section_title(ws_sum, r, 1, "APPLICATION-TO-OFFER RATE", end_col=6)
    r += 1
    for ci, hdr in enumerate(["Path", "Applied", "Offered", "Offer Rate"], 1):
        pill_cell(ws_sum, r, ci, hdr, CHARCOAL, font_color=WHITE, bold=True, size=10)
    ws_sum.row_dimensions[r].height = row_h

    _ALL_APP   = f'COUNTIFS({A_J},"Applied",{YC},{SC})'
    _ALL_OFF_A = f'COUNTIFS({A_T},"Offered",{YC},{SC})'
    _ATT_APP   = f'COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})'
    _ATT_OFF_A = f'COUNTIFS({A_G},"Registered and Attended",{A_T},"Offered",{YC},{SC})'
    _NS_APP    = f'COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})'
    _NS_OFF_A  = f'COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})'

    a2o_rows = [
        ("All Applicants", PILL_GREY, CHARCOAL,
         f'={_ALL_APP}', f'={_ALL_OFF_A}',
         f'=IF({_ALL_APP}=0,0,TEXT(ROUND({_ALL_OFF_A}/{_ALL_APP}*100,1),"0.0")&"%")'),
        ("Attended Path", PILL_GREEN, "1B7A3D",
         f'={_ATT_APP}', f'={_ATT_OFF_A}',
         f'=IF({_ATT_APP}=0,0,TEXT(ROUND({_ATT_OFF_A}/{_ATT_APP}*100,1),"0.0")&"%")'),
        ("No-Show Path", PILL_AMBER, "7D6608",
         f'={_NS_APP}', f'={_NS_OFF_A}',
         f'=IF({_NS_APP}=0,0,TEXT(ROUND({_NS_OFF_A}/{_NS_APP}*100,1),"0.0")&"%")'),
    ]
    for i, (label, fill, fc, app_f, off_f, rate_f) in enumerate(a2o_rows):
        rr = r + 1 + i
        pill_cell(ws_sum, rr, 1, label, fill, font_color=fc, size=10)
        pill_value(ws_sum, rr, 2, app_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 3, off_f, fill_hex=fill, font_color=fc, size=12)
        pill_value(ws_sum, rr, 4, rate_f, fill_hex=fill, font_color=fc, size=11)
        ws_sum.row_dimensions[rr].height = row_h

    a2o_end_row = r + len(a2o_rows)

    # ================================================================
    # SHEET 6 — Dashboard  (KPIs + Charts + Findings)
    # ================================================================
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.sheet_properties.tabColor = CHERRY
    ws_dash.sheet_view.showGridLines = False
    ws_dash.freeze_panes = "A6"

    # Light background fill (entire visible area)
    light_bg_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG,
                                fill_type="solid")
    for row in range(1, 201):
        for col in range(1, 13):
            ws_dash.cell(row=row, column=col).fill = light_bg_fill

    # La Trobe banner
    ltu_banner(ws_dash, "L")

    # Title — CHARCOAL bar, white text (clean professional look)
    ws_dash.merge_cells("A2:L3")
    title_cell = ws_dash.cell(row=2, column=1,
                              value="Webinar Pipeline Dashboard")
    title_cell.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=CHARCOAL, end_color=CHARCOAL,
                                  fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 13):
        ws_dash.cell(row=2, column=c).fill = PatternFill(
            start_color=CHARCOAL, end_color=CHARCOAL, fill_type="solid")
        ws_dash.cell(row=3, column=c).fill = PatternFill(
            start_color=CHARCOAL, end_color=CHARCOAL, fill_type="solid")

    # ─── Filter Dropdowns (row 4) — centrally positioned ─────
    # Year filter
    ws_dash.cell(row=4, column=4, value="Year:").font = Font(
        name="Calibri", size=11, bold=True, color=CHARCOAL)
    ws_dash.cell(row=4, column=4).alignment = Alignment(
        horizontal="right", vertical="center")
    yr_cell = ws_dash.cell(row=4, column=5, value="All")
    yr_cell.font = Font(name="Calibri", size=11, bold=True, color=CHERRY)
    yr_cell.alignment = Alignment(horizontal="center", vertical="center")
    yr_cell.border = Border(
        left=Side(style="thin", color=CHERRY),
        right=Side(style="thin", color=CHERRY),
        top=Side(style="thin", color=CHERRY),
        bottom=Side(style="thin", color=CHERRY))
    yr_cell.protection = CellProtection(locked=False)
    dv_yr_d = DataValidation(
        type="list", formula1='"All,2024,2025"', allow_blank=False)
    dv_yr_d.error = "Please select All, 2024, or 2025"
    dv_yr_d.errorTitle = "Invalid Year"
    ws_dash.add_data_validation(dv_yr_d)
    dv_yr_d.add(ws_dash["E4"])

    # Student Type filter
    ws_dash.cell(row=4, column=7, value="Student Type:").font = Font(
        name="Calibri", size=11, bold=True, color=CHARCOAL)
    ws_dash.cell(row=4, column=7).alignment = Alignment(
        horizontal="right", vertical="center")
    ws_dash.merge_cells("G4:H4")
    st_d_cell = ws_dash.cell(row=4, column=9, value="All")
    st_d_cell.font = Font(name="Calibri", size=11, bold=True, color=CHERRY)
    st_d_cell.alignment = Alignment(horizontal="center", vertical="center")
    st_d_cell.border = Border(
        left=Side(style="thin", color=CHERRY),
        right=Side(style="thin", color=CHERRY),
        top=Side(style="thin", color=CHERRY),
        bottom=Side(style="thin", color=CHERRY))
    st_d_cell.protection = CellProtection(locked=False)
    dv_st_d = DataValidation(
        type="list", formula1='"All,Domestic,International"', allow_blank=False)
    dv_st_d.error = "Please select All, Domestic, or International"
    dv_st_d.errorTitle = "Invalid Student Type"
    ws_dash.add_data_validation(dv_st_d)
    dv_st_d.add(ws_dash["I4"])

    ws_dash.row_dimensions[4].height = 28

    # Dynamic: shows filter state + zero warning
    dyn_dash5 = ws_dash.cell(row=5, column=1)
    dyn_dash5.value = (
        f'=IF({TOTAL_F}=0,'
        f'"No registrants match the current filters (Year="&{YF}&", Student="&{SF}&")",'
        f'"Showing "&TEXT({TOTAL_F},"#,##0")&" registrants")'
    )
    dyn_dash5.font = Font(name="Calibri", size=10, italic=True, color="999999")
    dyn_dash5.alignment = Alignment(horizontal="center", vertical="bottom")

    # ---- KPI Cards Row 1 (row 6-8): Registration → Application ----
    # Grouped so Attended sits next to Attended→Applied (Sophie's feedback)
    kpi_defs = [
        ("A", "B", "Total\nRegistrants",
         f'={TOTAL_F}', CHARCOAL),
        ("C", "D", "Attended",
         f'=COUNTIFS({A_G},"Registered and Attended",{YC},{SC})', LT_TEAL),
        ("E", "F", "Attended\n\u2192Applied %",
         f'=IF(COUNTIFS({A_G},"Registered and Attended",{YC},{SC})=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})/COUNTIFS({A_G},"Registered and Attended",{YC},{SC})*100,1),"0.0")&"%")',
         LT_TEAL),
        ("G", "H", "Attendance\nRate",
         f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_G},"Registered and Attended",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
         GREEN),
        ("I", "J", "Applied",
         f'=COUNTIFS({A_J},"Applied",{YC},{SC})', CHERRY),
        ("K", "L", "Conversion\nRate",
         f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_J},"Applied",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
         DARK_RED),
    ]

    def _kpi_card(ws, row_lbl, col1, col2, label, formula, color):
        """Render a single KPI card — white background, thin grey border, black/white font."""
        c1 = ord(col1) - 64
        c2 = ord(col2) - 64
        white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        grey_border = Border(
            left=Side(style="thin", color=WARM_GREY),
            right=Side(style="thin", color=WARM_GREY),
            top=Side(style="thin", color=WARM_GREY),
            bottom=Side(style="thin", color=WARM_GREY))
        # Label row — small bold, dark grey text on white
        ws.merge_cells(f"{col1}{row_lbl}:{col2}{row_lbl}")
        lbl = ws.cell(row=row_lbl, column=c1, value=label)
        lbl.font = Font(name="Calibri", size=9, bold=True, color=CHARCOAL)
        lbl.fill = white_fill
        lbl.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        ws.cell(row=row_lbl, column=c2).fill = white_fill
        # Value row — large bold black number
        ws.merge_cells(f"{col1}{row_lbl+1}:{col2}{row_lbl+2}")
        val = ws.cell(row=row_lbl + 1, column=c1)
        val.value = formula
        val.font = Font(name="Calibri", size=22, bold=True, color=CHARCOAL)
        val.fill = white_fill
        val.alignment = Alignment(horizontal="center", vertical="center")
        # Thin grey border around entire card
        for rr in range(row_lbl, row_lbl + 3):
            for cc in [c1, c2]:
                ws.cell(row=rr, column=cc).border = grey_border
                ws.cell(row=rr, column=cc).fill = white_fill

    for col1, col2, label, formula, color in kpi_defs:
        _kpi_card(ws_dash, 6, col1, col2, label, formula, color)

    ws_dash.row_dimensions[6].height = 35
    ws_dash.row_dimensions[7].height = 30
    ws_dash.row_dimensions[8].height = 30

    insight_q(ws_dash, 9, 1, "↳ Bottom row: what happens after they apply?")

    # ---- KPI Cards Row 2 (row 10-12): Offers → Enrolment ----
    kpi_defs2 = [
        ("A", "B", "Offered",
         f'=COUNTIFS({A_T},"Offered",{YC},{SC})', LT_TEAL),
        ("C", "D", "Offer Rate",
         f'=IF(COUNTIFS({A_J},"Applied",{YC},{SC})=0,0,TEXT(ROUND(COUNTIFS({A_T},"Offered",{YC},{SC})/COUNTIFS({A_J},"Applied",{YC},{SC})*100,1),"0.0")&"%")',
         GREEN),
        ("E", "F", "Enrolled",
         f'=COUNTIFS({A_U},"Enrolled",{YC},{SC})', CHERRY),
        ("G", "H", "Enrolment\nRate",
         f'=IF({TOTAL_F}=0,0,TEXT(ROUND(COUNTIFS({A_U},"Enrolled",{YC},{SC})/{TOTAL_F}*100,1),"0.0")&"%")',
         DARK_RED),
        ("I", "J", "No-Show\n→Applied %",
         f'=IF(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})=0,0,'
         f'TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})*100,1),"0.0")&"%")',
         CHERRY),
        ("K", "L", "No-Show\n→Offered %",
         f'=IF(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})=0,0,'
         f'TEXT(ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})*100,1),"0.0")&"%")',
         WARM_GREY),
    ]
    for col1, col2, label, formula, color in kpi_defs2:
        _kpi_card(ws_dash, 10, col1, col2, label, formula, color)

    ws_dash.row_dimensions[10].height = 35
    ws_dash.row_dimensions[11].height = 30
    ws_dash.row_dimensions[12].height = 30

    # ---- KPI Annotations (hover-able comments for managers) ----
    _ann = [
        # Row 1 KPIs — value cells at row 7 (reordered: related metrics grouped)
        (7, 1, "Total Registrants",
         "Count of all unique webinar registrations.\n"
         "Source: Analysis column D (Email — non-empty rows).\n"
         "Filtered by: Year (B2) and Student Type (E2)."),
        (7, 3, "Attended",
         "Registrants who actually attended the webinar session.\n"
         "Source: Analysis column G (Attendance_Category = 'Registered and Attended').\n"
         "Filtered by: Year and Student Type."),
        (7, 5, "Attended → Applied %",
         "Of those who attended, what % then applied?\n"
         "Formula: (Attended AND Applied) ÷ Attended × 100.\n"
         "Measures webinar effectiveness at driving applications."),
        (7, 7, "Attendance Rate",
         "Percentage of registrants who attended.\n"
         "Formula: Attended ÷ Total Registrants × 100.\n"
         "A higher rate indicates effective pre-event engagement."),
        (7, 9, "Applied",
         "Registrants who submitted a course application.\n"
         "Source: Analysis column J (Has_Applied = 'Applied').\n"
         "This is the primary conversion metric."),
        (7, 11, "Conversion Rate",
         "Percentage of ALL registrants who applied.\n"
         "Formula: Applied ÷ Total Registrants × 100.\n"
         "Measures overall funnel effectiveness."),
        # Row 2 KPIs — value cells at row 11
        (11, 1, "Offered",
         "Registrants who received a course offer.\n"
         "Source: Analysis column T (Has_Offer = 'Offered').\n"
         "Simulated at 70% of applicants for this prototype."),
        (11, 3, "Offer Rate",
         "Percentage of applicants who received an offer.\n"
         "Formula: Offered ÷ Applied × 100.\n"
         "Indicates application quality and course capacity."),
        (11, 5, "Enrolled",
         "Registrants who completed enrolment.\n"
         "Source: Analysis column U (Has_Enrolled = 'Enrolled').\n"
         "Simulated at 60% of offered for this prototype."),
        (11, 7, "Enrolment Rate",
         "Percentage of ALL registrants who enrolled.\n"
         "Formula: Enrolled ÷ Total Registrants × 100.\n"
         "The bottom-line metric for webinar ROI."),
        (11, 9, "No-Show → Applied %",
         "Of those who did NOT attend, what % still applied?\n"
         "Formula: (Not Attended AND Applied) ÷ Not Attended × 100.\n"
         "Shows hidden value — some registrants convert without attending."),
        (11, 11, "No-Show → Offered %",
         "Of those who did NOT attend, what % received an offer?\n"
         "Formula: (Not Attended AND Offered) ÷ Not Attended × 100.\n"
         "Measures pipeline viability of non-attendees."),
    ]
    for ann_row, ann_col, ann_title, ann_text in _ann:
        cell = ws_dash.cell(row=ann_row, column=ann_col)
        cell.comment = Comment(f"{ann_title}\n\n{ann_text}", "La Trobe Analytics")

    # Column widths — narrow gutters (A, G) + wider content columns
    for c in range(1, 13):
        if c in (1, 7):  # gutter columns
            ws_dash.column_dimensions[get_column_letter(c)].width = 3
        else:
            ws_dash.column_dimensions[get_column_letter(c)].width = 14

    # Dynamic: warns when filtered total is zero
    dyn_dash13 = ws_dash.cell(row=13, column=1)
    dyn_dash13.value = (
        f'=IF({TOTAL_F}=0,'
        f'"⚠ No data for this filter — try Year=All and Student=All",'
        f'"↳ Charts below show breakdown · Filtered: "'
        f'&IF({YF}="All","All years",{YF})'
        f'&" · "&IF({SF}="All","All students",{SF}))'
    )
    dyn_dash13.font = Font(name="Calibri", size=10, italic=True, color="999999")
    dyn_dash13.alignment = Alignment(horizontal="left", vertical="bottom")

    # ---- Charts ----
    # The charts reference the Summary pill cells for their data.
    # We need simple label+value tables for chart references.
    # We'll put hidden chart-data in columns G-H of Summary (right side)
    # so we don't mess up the pill layout.

    # Chart data area in Summary (cols G-J, hidden from main view)
    ws_sum.column_dimensions["G"].width = 24
    ws_sum.column_dimensions["H"].width = 12
    ws_sum.column_dimensions["I"].width = 12
    ws_sum.column_dimensions["J"].width = 12

    cd_font = Font(name="Calibri", size=9, color=WARM_GREY)

    def chart_data_row(ws, r, col_label, col_val, label, formula):
        ws.cell(row=r, column=col_label, value=label).font = cd_font
        ws.cell(row=r, column=col_val).value = formula
        ws.cell(row=r, column=col_val).font = cd_font

    # Attendance chart data (G2:H3)
    chart_data_row(ws_sum, 2, 7, 8, "Attended",
                   f'=COUNTIFS({A_G},"Registered and Attended",{YC},{SC})')
    chart_data_row(ws_sum, 3, 7, 8, "Not Attended",
                   f'=COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})')

    # Application chart data (G5:H6)
    chart_data_row(ws_sum, 5, 7, 8, "Applied",
                   f'=COUNTIFS({A_J},"Applied",{YC},{SC})')
    chart_data_row(ws_sum, 6, 7, 8, "Did Not Apply",
                   f'=COUNTIFS({A_J},"Did Not Apply",{YC},{SC})')

    # Cross-tab chart data (G8:I10) — after application data
    ws_sum.cell(row=18, column=8, value="Applied").font = cd_font
    ws_sum.cell(row=18, column=9, value="Did Not Apply").font = cd_font
    ws_sum.cell(row=19, column=7, value="Attended").font = cd_font
    ws_sum.cell(row=19, column=8).value = \
        f'=COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})'
    ws_sum.cell(row=19, column=9).value = \
        f'=COUNTIFS({A_G},"Registered and Attended",{A_J},"Did Not Apply",{YC},{SC})'
    ws_sum.cell(row=20, column=7, value="Not Attended").font = cd_font
    ws_sum.cell(row=20, column=8).value = \
        f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})'
    ws_sum.cell(row=20, column=9).value = \
        f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Did Not Apply",{YC},{SC})'

    # Discovery channel chart data — header row + data
    ws_sum.cell(row=21, column=8, value="Total").font = cd_font
    ws_sum.cell(row=21, column=9, value="Applied").font = cd_font
    for i, ch in enumerate(HOW_FOUND):
        rr = 22 + i
        ws_sum.cell(row=rr, column=7, value=ch).font = cd_font
        ws_sum.cell(row=rr, column=8).value = f'=COUNTIFS({A_P},"{ch}",{YC},{SC})'
        ws_sum.cell(row=rr, column=9).value = \
            f'=COUNTIFS({A_P},"{ch}",{A_J},"Applied",{YC},{SC})'
    disc_chart_end = 22 + len(HOW_FOUND) - 1

    # Age range chart data — header row + data
    age_header_row = disc_chart_end + 2
    ws_sum.cell(row=age_header_row, column=8, value="Total").font = cd_font
    ws_sum.cell(row=age_header_row, column=9, value="Applied").font = cd_font
    age_chart_start = age_header_row + 1
    for i, ar in enumerate(AGE_RANGES):
        rr = age_chart_start + i
        ws_sum.cell(row=rr, column=7, value=ar).font = cd_font
        ws_sum.cell(row=rr, column=8).value = f'=COUNTIFS({A_N},"{ar}",{YC},{SC})'
        ws_sum.cell(row=rr, column=9).value = \
            f'=COUNTIFS({A_N},"{ar}",{A_J},"Applied",{YC},{SC})'
    age_chart_end = age_chart_start + len(AGE_RANGES) - 1

    # Campus chart data — header row + data
    campus_header_row = age_chart_end + 2
    ws_sum.cell(row=campus_header_row, column=8, value="Total").font = cd_font
    ws_sum.cell(row=campus_header_row, column=9, value="Applied").font = cd_font
    campus_chart_start = campus_header_row + 1
    for i, cp in enumerate(CAMPUSES):
        rr = campus_chart_start + i
        ws_sum.cell(row=rr, column=7, value=cp).font = cd_font
        ws_sum.cell(row=rr, column=8).value = f'=COUNTIFS({A_M},"{cp}",{YC},{SC})'
        ws_sum.cell(row=rr, column=9).value = \
            f'=COUNTIFS({A_M},"{cp}",{A_J},"Applied",{YC},{SC})'
    campus_chart_end = campus_chart_start + len(CAMPUSES) - 1

    # Enrolment Funnel chart data (after campus)
    funnel_header_row = campus_chart_end + 2
    ws_sum.cell(row=funnel_header_row, column=7, value="Stage").font = cd_font
    ws_sum.cell(row=funnel_header_row, column=8, value="Count").font = cd_font
    funnel_chart_start = funnel_header_row + 1
    funnel_chart_data = [
        ("Enrolled", f'=COUNTIFS({A_U},"Enrolled",{YC},{SC})'),
        ("Offered", f'=COUNTIFS({A_T},"Offered",{YC},{SC})'),
        ("Applied", f'=COUNTIFS({A_J},"Applied",{YC},{SC})'),
        ("Attended", f'=COUNTIFS({A_G},"Registered and Attended",{YC},{SC})'),
        ("Registered", f'={TOTAL_F}'),
    ]
    for i, (label, formula) in enumerate(funnel_chart_data):
        rr = funnel_chart_start + i
        ws_sum.cell(row=rr, column=7, value=label).font = cd_font
        ws_sum.cell(row=rr, column=8).value = formula
        ws_sum.cell(row=rr, column=8).font = cd_font
    funnel_chart_end = funnel_chart_start + len(funnel_chart_data) - 1

    # School Pipeline chart data (after funnel) — 3 series per school
    school_header_row = funnel_chart_end + 2
    ws_sum.cell(row=school_header_row, column=8, value="Registered").font = cd_font
    ws_sum.cell(row=school_header_row, column=9, value="Applied").font = cd_font
    ws_sum.cell(row=school_header_row, column=10, value="Enrolled").font = cd_font
    school_chart_start = school_header_row + 1
    all_schools_chart = SCHOOLS + ["Unknown"]
    for i, school in enumerate(all_schools_chart):
        rr = school_chart_start + i
        short = school if len(school) <= 25 else school[:23] + ".."
        ws_sum.cell(row=rr, column=7, value=short).font = cd_font
        ws_sum.cell(row=rr, column=8).value = \
            f'=COUNTIFS({A_S},"{school}",{YC},{SC})'
        ws_sum.cell(row=rr, column=9).value = \
            f'=COUNTIFS({A_S},"{school}",{A_J},"Applied",{YC},{SC})'
        ws_sum.cell(row=rr, column=10).value = \
            f'=COUNTIFS({A_S},"{school}",{A_U},"Enrolled",{YC},{SC})'
    school_chart_end = school_chart_start + len(all_schools_chart) - 1

    # Domestic vs International chart data (after school pipeline)
    student_header_row = school_chart_end + 2
    ws_sum.cell(row=student_header_row, column=8, value="Registered").font = cd_font
    ws_sum.cell(row=student_header_row, column=9, value="Applied").font = cd_font
    ws_sum.cell(row=student_header_row, column=10, value="Enrolled").font = cd_font
    student_chart_start = student_header_row + 1
    for i, stype in enumerate(["Domestic", "International"]):
        rr = student_chart_start + i
        ws_sum.cell(row=rr, column=7, value=stype).font = cd_font
        ws_sum.cell(row=rr, column=8).value = \
            f'=COUNTIFS({A_W},"{stype}",{YC},{SC})'
        ws_sum.cell(row=rr, column=9).value = \
            f'=COUNTIFS({A_W},"{stype}",{A_J},"Applied",{YC},{SC})'
        ws_sum.cell(row=rr, column=10).value = \
            f'=COUNTIFS({A_W},"{stype}",{A_U},"Enrolled",{YC},{SC})'
    student_chart_end = student_chart_start + 1  # 2 rows

    # No-Show Conversion chart data (after student type)
    noshow_header_row = student_chart_end + 2
    ws_sum.cell(row=noshow_header_row, column=7, value="Stage").font = cd_font
    ws_sum.cell(row=noshow_header_row, column=8, value="Count").font = cd_font
    noshow_chart_start = noshow_header_row + 1
    noshow_chart_data = [
        ("Did Not Attend",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})'),
        ("Still Applied",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})'),
        ("Received Offer",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})'),
        ("Enrolled",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_U},"Enrolled",{YC},{SC})'),
    ]
    for i, (label, formula) in enumerate(noshow_chart_data):
        rr = noshow_chart_start + i
        ws_sum.cell(row=rr, column=7, value=label).font = cd_font
        ws_sum.cell(row=rr, column=8).value = formula
        ws_sum.cell(row=rr, column=8).font = cd_font
    noshow_chart_end = noshow_chart_start + len(noshow_chart_data) - 1

    # Application-to-Offer chart data (after no-show conversion)
    a2o_chart_header = noshow_chart_end + 2
    ws_sum.cell(row=a2o_chart_header, column=7, value="Path").font = cd_font
    ws_sum.cell(row=a2o_chart_header, column=8, value="Applied").font = cd_font
    ws_sum.cell(row=a2o_chart_header, column=9, value="Offered").font = cd_font
    a2o_chart_start = a2o_chart_header + 1
    for i, (label, app_f, off_f) in enumerate([
        ("Attended",
         f'=COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})',
         f'=COUNTIFS({A_G},"Registered and Attended",{A_T},"Offered",{YC},{SC})'),
        ("No-Show",
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})',
         f'=COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})'),
    ]):
        rr = a2o_chart_start + i
        ws_sum.cell(row=rr, column=7, value=label).font = cd_font
        ws_sum.cell(row=rr, column=8).value = app_f
        ws_sum.cell(row=rr, column=8).font = cd_font
        ws_sum.cell(row=rr, column=9).value = off_f
        ws_sum.cell(row=rr, column=9).font = cd_font
    a2o_chart_end = a2o_chart_start + 1

    # ---- Now create the charts referencing the chart data area ----

    # Chart 1: Attendance Donut
    c1 = DoughnutChart()
    c1.title = "Attendance Breakdown"
    c1.style = 10
    c1.add_data(Reference(ws_sum, min_col=8, min_row=2, max_row=3))
    c1.set_categories(Reference(ws_sum, min_col=7, min_row=2, max_row=3))
    c1.width = 13; c1.height = 8
    s = c1.series[0]
    p0 = DataPoint(idx=0); p0.graphicalProperties.solidFill = "3498DB"
    p1 = DataPoint(idx=1); p1.graphicalProperties.solidFill = "95A5A6"
    s.data_points = [p0, p1]
    s.dLbls = DataLabelList(); s.dLbls.showPercent = True; s.dLbls.showCatName = True; s.dLbls.showVal = True
    ws_dash.add_chart(c1, "A46")

    # Chart 2: Application Donut
    c2 = DoughnutChart()
    c2.title = "Application Status"
    c2.style = 10
    c2.add_data(Reference(ws_sum, min_col=8, min_row=5, max_row=6))
    c2.set_categories(Reference(ws_sum, min_col=7, min_row=5, max_row=6))
    c2.width = 13; c2.height = 8
    s = c2.series[0]
    p0 = DataPoint(idx=0); p0.graphicalProperties.solidFill = "2C3E50"
    p1 = DataPoint(idx=1); p1.graphicalProperties.solidFill = "BDC3C7"
    s.data_points = [p0, p1]
    s.dLbls = DataLabelList(); s.dLbls.showPercent = True; s.dLbls.showCatName = True; s.dLbls.showVal = True
    ws_dash.add_chart(c2, "G46")

    # Chart 3: Cross-tab — side-by-side bars (shifted for 5-stage pipeline)
    c5 = BarChart()
    c5.type = "col"; c5.grouping = "clustered"
    c5.title = "Attendance vs Application (Cross-Tab)"; c5.style = 10
    c5.add_data(Reference(ws_sum, min_col=8, max_col=9, min_row=18, max_row=20),
                titles_from_data=True)
    c5.set_categories(Reference(ws_sum, min_col=7, min_row=19, max_row=20))
    c5.width = 13; c5.height = 8
    c5.series[0].graphicalProperties.solidFill = "3498DB"
    c5.series[1].graphicalProperties.solidFill = "2C3E50"
    c5.y_axis.title = "Registrants"
    c5.dataLabels = DataLabelList(); c5.dataLabels.showVal = True
    ws_dash.add_chart(c5, "A62")

    # Chart 6: Discovery Channel — side-by-side bars
    c6 = BarChart()
    c6.type = "col"; c6.grouping = "clustered"
    c6.title = "How They Found Us"; c6.style = 10
    c6.add_data(Reference(ws_sum, min_col=8, max_col=9, min_row=21,
                          max_row=disc_chart_end), titles_from_data=True)
    c6.set_categories(Reference(ws_sum, min_col=7, min_row=22,
                                max_row=disc_chart_end))
    c6.width = 13; c6.height = 8
    c6.series[0].graphicalProperties.solidFill = "1ABC9C"
    c6.series[1].graphicalProperties.solidFill = "95A5A6"
    c6.y_axis.title = "Count"
    c6.dataLabels = DataLabelList(); c6.dataLabels.showVal = True
    ws_dash.add_chart(c6, "G62")

    # Chart 7: Age Range — side-by-side bars
    c7 = BarChart()
    c7.type = "col"; c7.grouping = "clustered"
    c7.title = "Age Range Breakdown"; c7.style = 10
    c7.add_data(Reference(ws_sum, min_col=8, max_col=9,
                          min_row=age_header_row, max_row=age_chart_end),
                titles_from_data=True)
    c7.set_categories(Reference(ws_sum, min_col=7,
                                min_row=age_chart_start, max_row=age_chart_end))
    c7.width = 13; c7.height = 8
    c7.series[0].graphicalProperties.solidFill = "2C3E50"
    c7.series[1].graphicalProperties.solidFill = "3498DB"
    c7.y_axis.title = "Count"
    c7.dataLabels = DataLabelList(); c7.dataLabels.showVal = True
    ws_dash.add_chart(c7, "A78")

    # Chart 8: Campus — side-by-side bars
    c8 = BarChart()
    c8.type = "col"; c8.grouping = "clustered"
    c8.title = "Campus Interest"; c8.style = 10
    c8.add_data(Reference(ws_sum, min_col=8, max_col=9,
                          min_row=campus_header_row, max_row=campus_chart_end),
                titles_from_data=True)
    c8.set_categories(Reference(ws_sum, min_col=7,
                                min_row=campus_chart_start, max_row=campus_chart_end))
    c8.width = 13; c8.height = 8
    c8.series[0].graphicalProperties.solidFill = "1ABC9C"
    c8.series[1].graphicalProperties.solidFill = "95A5A6"
    c8.y_axis.title = "Count"
    c8.dataLabels = DataLabelList(); c8.dataLabels.showVal = True
    ws_dash.add_chart(c8, "G78")

    # Chart 9: Enrolment Funnel — horizontal bar (Registered at top → Enrolled at bottom)
    c9 = BarChart()
    c9.type = "bar"; c9.title = "Enrolment Funnel"; c9.style = 10
    c9.add_data(Reference(ws_sum, min_col=8,
                          min_row=funnel_chart_start, max_row=funnel_chart_end))
    c9.set_categories(Reference(ws_sum, min_col=7,
                                min_row=funnel_chart_start, max_row=funnel_chart_end))
    c9.width = 13; c9.height = 8
    c9.series[0].graphicalProperties.solidFill = "3498DB"
    c9.y_axis.title = "Stage"
    c9.x_axis.title = "Count"
    c9.dataLabels = DataLabelList(); c9.dataLabels.showVal = True
    ws_dash.add_chart(c9, "A94")

    # Chart 10: School Pipeline — clustered bar per school
    c10 = BarChart()
    c10.type = "col"; c10.grouping = "clustered"
    c10.title = "School Pipeline"; c10.style = 10
    c10.add_data(Reference(ws_sum, min_col=8, max_col=10,
                           min_row=school_header_row, max_row=school_chart_end),
                 titles_from_data=True)
    c10.set_categories(Reference(ws_sum, min_col=7,
                                 min_row=school_chart_start,
                                 max_row=school_chart_end))
    c10.width = 13; c10.height = 8
    c10.series[0].graphicalProperties.solidFill = "2C3E50"
    c10.series[1].graphicalProperties.solidFill = "3498DB"
    c10.series[2].graphicalProperties.solidFill = "1ABC9C"
    c10.y_axis.title = "Count"
    c10.dataLabels = DataLabelList(); c10.dataLabels.showVal = True
    ws_dash.add_chart(c10, "G94")

    # Chart 11: Domestic vs International — clustered bar
    c11 = BarChart()
    c11.type = "col"; c11.grouping = "clustered"
    c11.title = "Domestic vs International Pipeline"; c11.style = 10
    c11.add_data(Reference(ws_sum, min_col=8, max_col=10,
                           min_row=student_header_row, max_row=student_chart_end),
                 titles_from_data=True)
    c11.set_categories(Reference(ws_sum, min_col=7,
                                 min_row=student_chart_start,
                                 max_row=student_chart_end))
    c11.width = 13; c11.height = 8
    c11.series[0].graphicalProperties.solidFill = "2C3E50"
    c11.series[1].graphicalProperties.solidFill = "3498DB"
    c11.series[2].graphicalProperties.solidFill = "1ABC9C"
    c11.y_axis.title = "Count"
    c11.dataLabels = DataLabelList(); c11.dataLabels.showVal = True
    ws_dash.add_chart(c11, "A110")

    # Chart 12: No-Show Conversion — horizontal bar
    c12 = BarChart()
    c12.type = "bar"; c12.title = "No-Show Conversion"; c12.style = 10
    c12.add_data(Reference(ws_sum, min_col=8,
                           min_row=noshow_chart_start, max_row=noshow_chart_end))
    c12.set_categories(Reference(ws_sum, min_col=7,
                                 min_row=noshow_chart_start, max_row=noshow_chart_end))
    c12.width = 13; c12.height = 8
    c12.series[0].graphicalProperties.solidFill = "95A5A6"
    c12.y_axis.title = "Stage"
    c12.x_axis.title = "Count"
    c12.dataLabels = DataLabelList(); c12.dataLabels.showVal = True
    ws_dash.add_chart(c12, "G110")

    # Chart 13: Application-to-Offer Rate — clustered bar
    c15 = BarChart()
    c15.type = "col"; c15.grouping = "clustered"
    c15.title = "Application-to-Offer Rate by Path"; c15.style = 10
    c15.add_data(Reference(ws_sum, min_col=8, max_col=9,
                           min_row=a2o_chart_header, max_row=a2o_chart_end),
                 titles_from_data=True)
    c15.set_categories(Reference(ws_sum, min_col=7,
                                 min_row=a2o_chart_start, max_row=a2o_chart_end))
    c15.width = 13; c15.height = 8
    c15.series[0].graphicalProperties.solidFill = "3498DB"
    c15.series[1].graphicalProperties.solidFill = "2C3E50"
    c15.y_axis.title = "Count"
    c15.dataLabels = DataLabelList(); c15.dataLabels.showVal = True
    ws_dash.add_chart(c15, "A126")

    # ---- Section dividers between chart groups ----
    divider_font = Font(name="Calibri", size=10, bold=True, color=WARM_GREY)
    divider_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG,
                               fill_type="solid")
    divider_border = Border(
        bottom=Side(style="thin", color=WARM_GREY))
    chart_sections = [
        (45, "OVERVIEW"),
        (61, "DEMOGRAPHICS & DISCOVERY"),
        (77, "CAMPUS & AGE"),
        (93, "PIPELINE & SCHOOLS"),
        (109, "STUDENT MIX & CONVERSION"),
        (125, "OFFER ANALYSIS"),
    ]
    for div_row, div_title in chart_sections:
        ws_dash.merge_cells(f"B{div_row}:L{div_row}")
        dc = ws_dash.cell(row=div_row, column=2, value=div_title)
        dc.font = divider_font
        dc.fill = divider_fill
        dc.border = divider_border
        dc.alignment = Alignment(horizontal="left", vertical="bottom")
        for cc in range(3, 13):
            ws_dash.cell(row=div_row, column=cc).border = divider_border

    # ---- Data bars on school pipeline values (Summary sheet) ----
    school_reg_range = f"H{school_chart_start}:H{school_chart_end}"
    school_app_range = f"I{school_chart_start}:I{school_chart_end}"
    school_enr_range = f"J{school_chart_start}:J{school_chart_end}"
    for rng in [school_reg_range, school_app_range, school_enr_range]:
        ws_sum.conditional_formatting.add(
            rng,
            DataBarRule(start_type="min", end_type="max",
                        color="3498DB"))

    # ---- KEY FINDINGS (row 14 — prominent position after KPIs) ----
    fr = 14
    ws_dash.merge_cells(f"A{fr}:L{fr}")
    ft = ws_dash.cell(row=fr, column=1, value="KEY FINDINGS")
    ft.font = Font(name="Calibri", size=14, bold=True, color=CHARCOAL)
    ft.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ft.alignment = Alignment(horizontal="left", vertical="center")
    ft.border = Border(bottom=Side(style="medium", color=CHARCOAL))
    for c in range(2, 13):
        ws_dash.cell(row=fr, column=c).border = Border(
            bottom=Side(style="medium", color=CHARCOAL))

    fs = fr + 2
    ffont = Font(name="Calibri", size=11, color=CHARCOAL)
    bfont = Font(name="Calibri", size=11, bold=True, color=CHERRY)

    findings = [
        # #1 — Overall Conversion (year-filtered)
        (f'=CONCATENATE("Overall Conversion: ",'
         f'TEXT(IF({TOTAL_F}=0,0,ROUND(COUNTIFS({A_J},"Applied",{YC},{SC})/{TOTAL_F}*100,1)),"0.0"),'
         f'"% of webinar registrants applied. ",'
         f'IF(COUNTIFS({A_J},"Applied",{YC},{SC})/MAX({TOTAL_F},1)>0.2,'
         f'"Strong conversion.","Room to improve follow-up."))'),

        # #2 — Attendance Impact (year-filtered)
        (f'=CONCATENATE("Attendance Impact: ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Attended",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Attended",{YC},{SC})*100,1)),"0.0"),'
         f'"% of attendees applied vs ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})*100,1)),"0.0"),'
         f'"% of no-shows.")'),

        # #3 — Full Pipeline (year-filtered)
        (f'=CONCATENATE("Pipeline: ",'
         f'TEXT({TOTAL_F},"0")," registrants \u2192 ",'
         f'TEXT(COUNTIFS({A_J},"Applied",{YC},{SC}),"0")," applicants \u2192 ",'
         f'TEXT(COUNTIFS({A_T},"Offered",{YC},{SC}),"0")," offered \u2192 ",'
         f'TEXT(COUNTIFS({A_U},"Enrolled",{YC},{SC}),"0")," enrolled.")'),

        # #5 — Offer Rate (new)
        (f'=CONCATENATE("Offer Rate: ",'
         f'TEXT(IF(COUNTIFS({A_J},"Applied",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_T},"Offered",{YC},{SC})/COUNTIFS({A_J},"Applied",{YC},{SC})*100,1)),"0.0"),'
         f'"% of applicants received offers. ",'
         f'IF(COUNTIFS({A_T},"Offered",{YC},{SC})/MAX(COUNTIFS({A_J},"Applied",{YC},{SC}),1)>0.5,'
         f'"Healthy offer rate.","May need admissions review."))'),

        # #6 — Enrolment Yield (new)
        (f'=CONCATENATE("Enrolment Yield: ",'
         f'TEXT(COUNTIFS({A_U},"Enrolled",{YC},{SC}),"0")," of ",'
         f'TEXT({TOTAL_F},"0")," registrants enrolled (",'
         f'TEXT(IF({TOTAL_F}=0,0,ROUND(COUNTIFS({A_U},"Enrolled",{YC},{SC})/{TOTAL_F}*100,1)),"0.0"),'
         f'"% yield). ",'
         f'IF(COUNTIFS({A_T},"Offered",{YC},{SC})=0,"No offers yet.",'
         f'CONCATENATE(TEXT(IF(COUNTIFS({A_T},"Offered",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_U},"Enrolled",{YC},{SC})/COUNTIFS({A_T},"Offered",{YC},{SC})*100,1)),"0.0"),'
         f'"% of offers converted to enrolment.")))'),
        # #7 — Student Mix (Domestic vs International)
        (f'=CONCATENATE("Student Mix: ",'
         f'TEXT(COUNTIFS({A_W},"Domestic",{YC},{SC}),"0")," domestic (",'
         f'TEXT(IF({TOTAL_F}=0,0,ROUND(COUNTIFS({A_W},"Domestic",{YC},{SC})/{TOTAL_F}*100,1)),"0.0"),'
         f'"%) and ",'
         f'TEXT(COUNTIFS({A_W},"International",{YC},{SC}),"0")," international (",'
         f'TEXT(IF({TOTAL_F}=0,0,ROUND(COUNTIFS({A_W},"International",{YC},{SC})/{TOTAL_F}*100,1)),"0.0"),'
         f'"%) registrants.")'),

        # #8 — Pipeline Conversion (stage-to-stage rates — manager's diagram)
        (f'=CONCATENATE("Funnel Conversion: ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Attended",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Attended",{YC},{SC})*100,1)),"0.0"),'
         f'"% of attendees applied, ",'
         f'TEXT(IF(COUNTIFS({A_J},"Applied",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_T},"Offered",{YC},{SC})/'
         f'COUNTIFS({A_J},"Applied",{YC},{SC})*100,1)),"0.0"),'
         f'"% of applicants offered, ",'
         f'TEXT(IF(COUNTIFS({A_T},"Offered",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_U},"Enrolled",{YC},{SC})/'
         f'COUNTIFS({A_T},"Offered",{YC},{SC})*100,1)),"0.0"),'
         f'"% of offered enrolled.")'),

        # #9 — No-Show Conversion (manager's 3rd question)
        (f'=CONCATENATE("No-Show Value: ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})*100,1)),"0.0"),'
         f'"% of non-attendees still applied, ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{YC},{SC})*100,1)),"0.0"),'
         f'"% received offers.")'),

        # #10 — Domestic vs International Conversion Rates
        (f'=CONCATENATE("Conversion by Type: Domestic ",'
         f'TEXT(IF(COUNTIFS({A_W},"Domestic",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_W},"Domestic",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_W},"Domestic",{YC},{SC})*100,1)),"0.0"),'
         f'"% applied, International ",'
         f'TEXT(IF(COUNTIFS({A_W},"International",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_W},"International",{A_J},"Applied",{YC},{SC})/'
         f'COUNTIFS({A_W},"International",{YC},{SC})*100,1)),"0.0"),'
         f'"% applied.")'),

        # #12 — Application Quality (are webinar applicants strong candidates?)
        (f'=CONCATENATE("Application Quality: ",'
         f'TEXT(IF(COUNTIFS({A_J},"Applied",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_T},"Offered",{YC},{SC})/COUNTIFS({A_J},"Applied",{YC},{SC})*100,1)),"0.0"),'
         f'"% offer rate overall. Attendees: ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Attended",{A_T},"Offered",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Attended",{A_J},"Applied",{YC},{SC})*100,1)),"0.0"),'
         f'"%, No-Shows: ",'
         f'TEXT(IF(COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})=0,0,'
         f'ROUND(COUNTIFS({A_G},"Registered and Not Attended",{A_T},"Offered",{YC},{SC})/'
         f'COUNTIFS({A_G},"Registered and Not Attended",{A_J},"Applied",{YC},{SC})*100,1)),"0.0"),'
         f'"% offer rate.")'),

        # #14 — Top-Performing School (highest enrolled count from chart data)
        (f'=CONCATENATE("Top School: ",'
         f'INDEX(Summary!$G${school_chart_start}:$G${school_chart_end},'
         f'MATCH(MAX(Summary!$J${school_chart_start}:$J${school_chart_end}),'
         f'Summary!$J${school_chart_start}:$J${school_chart_end},0)),'
         f'" leads with ",'
         f'TEXT(MAX(Summary!$J${school_chart_start}:$J${school_chart_end}),"0"),'
         f'" enrolled students.")'),

        # #15 — Year-on-Year (compare 2024 vs 2025 — uses SC but NOT YC)
        (f'=CONCATENATE("Year-on-Year: ",'
         f'TEXT(COUNTIFS({A_J},"Applied",{A_R},"2024",{SC}),"0"),'
         f'" applied in 2024 vs ",'
         f'TEXT(COUNTIFS({A_J},"Applied",{A_R},"2025",{SC}),"0"),'
         f'" in 2025. Enrolled: ",'
         f'TEXT(COUNTIFS({A_U},"Enrolled",{A_R},"2024",{SC}),"0"),'
         f'" → ",'
         f'TEXT(COUNTIFS({A_U},"Enrolled",{A_R},"2025",{SC}),"0"),'
         f'".")'),
    ]

    for i, formula in enumerate(findings):
        r = fs + i * 2
        ws_dash.cell(row=r, column=1, value=f"#{i+1}").font = bfont
        ws_dash.merge_cells(f"B{r}:L{r}")
        ws_dash.cell(row=r, column=2).value = formula
        ws_dash.cell(row=r, column=2).font = ffont
        ws_dash.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws_dash.row_dimensions[r].height = 35

    # Footer
    ftr = fs + len(findings) * 2 + 2
    ws_dash.merge_cells(f"A{ftr}:L{ftr}")
    ws_dash.cell(row=ftr, column=1,
                 value="La Trobe University \u2014 Dashboard auto-updates when data changes").font = \
        Font(name="Calibri", size=9, italic=True, color=WARM_GREY)

    # ---- Dashboard Footer ----
    footer_row = 142
    ws_dash.merge_cells(f"A{footer_row}:L{footer_row}")
    ftr = ws_dash.cell(row=footer_row, column=1,
                       value="La Trobe University \u00b7 Postgraduate Webinar Pipeline Analysis \u00b7 V9")
    ftr.font = Font(name="Calibri", size=9, italic=True, color=WARM_GREY)
    ftr.alignment = Alignment(horizontal="center", vertical="center")
    ftr.border = Border(top=Side(style="thin", color=WARM_GREY))
    for cc in range(2, 13):
        ws_dash.cell(row=footer_row, column=cc).border = Border(
            top=Side(style="thin", color=WARM_GREY))

    # ================================================================
    # Protect formula sheets
    # ================================================================
    ws_anal.protection.sheet = True
    ws_anal.protection.password = "view"
    ws_sum.protection.sheet = True
    ws_sum.protection.password = "view"
    ws_dash.protection.sheet = True
    ws_dash.protection.password = "view"

    # ================================================================
    # Reorder tabs: Instructions → Dashboard → Summary → Analysis
    #               → Webinar data → Flexi PG → SchoolMap (hidden)
    # ================================================================
    desired_order = [
        "Instructions",
        "Dashboard",
        "Summary",
        "Analysis",
        "Webinar data 2024_25",
        "Flexi PG 2024_25",
        "SchoolMap",
    ]
    for target_idx, name in enumerate(desired_order):
        current_idx = wb.sheetnames.index(name)
        if current_idx != target_idx:
            wb.move_sheet(name, offset=target_idx - current_idx)

    # Set Dashboard as the active (visible) sheet on open
    wb.active = wb.sheetnames.index("Dashboard")

    # ================================================================
    # Save
    # ================================================================
    wb.save(OUTPUT_FILE)
    print(f"\n  Workbook saved: {OUTPUT_FILE}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  La Trobe University brand theme applied")
    print(f"  Apple-style pills on Summary tab")
    print(f"  Year filter on Dashboard!E4 (All / 2024 / 2025)")
    print(f"  Student type filter on Dashboard!I4 (All / Domestic / International)")
    print(f"  School breakdown with per-school pipeline")
    print(f"  Domestic vs International breakdown")
    print(f"  Application-to-Offer Rate")
    print(f"  11 charts · 12 findings · 12 KPI cards")
    print(f"  Formulas cover rows 3\u2013{MAX_DATA_ROW + 2}")
    print(f"\n  Copy this file to your work computer \u2014 no scripts needed.\n")


if __name__ == "__main__":
    build()
